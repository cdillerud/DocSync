"""Deterministic parser for CanPack supplier/manufacturer sales-order schedule XLSX files."""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook

from ..models import (
    DocumentType,
    NormalizedInboundOrder,
    NormalizedRelease,
    OrderCustomer,
    OrderDocument,
    OrderSource,
    OrderValidation,
    ProposedAction,
)
from ..utils import clean_text, excel_serial_to_datetime, sha256_file


class CanPackXlsxParser:
    SOURCE_FORMAT = "CANPACK_XLSX"

    HEADER_ALIASES: Dict[str, List[str]] = {
        "po": ["po#", "po", "purchase order", "purchase order #"],
        "description": ["customer design description", "design description"],
        "customer_item": ["customer material number", "customer material no", "material number"],
        "quantity": ["call-off quantity", "call off quantity", "quantity"],
        "uom": ["unit of measurement (uom)", "unit of measurement", "uom"],
        "plant": ["delivering plant", "plant"],
        "pickup": ["customer expected pick up date and time", "expected pickup", "expected pick up"],
        "receive": ["customer requested receive by date", "requested receive by date", "receive by date"],
    }

    def __init__(self, sheet_name: Optional[str] = None):
        self.sheet_name = sheet_name

    @staticmethod
    def _normalized_header(value) -> str:
        return " ".join(str(value or "").strip().lower().split())

    def _find_header_row(self, ws) -> tuple[int, Dict[str, int]]:
        for row_idx in range(1, min(ws.max_row, 25) + 1):
            row_values = {
                col_idx: self._normalized_header(ws.cell(row_idx, col_idx).value)
                for col_idx in range(1, ws.max_column + 1)
            }
            mapping: Dict[str, int] = {}
            for key, aliases in self.HEADER_ALIASES.items():
                alias_set = {self._normalized_header(a) for a in aliases}
                for col_idx, value in row_values.items():
                    if value in alias_set:
                        mapping[key] = col_idx
                        break
            if "po" in mapping and "customer_item" in mapping and "quantity" in mapping:
                return row_idx, mapping
        raise ValueError("Could not identify CanPack header row")

    def parse(self, path: str | Path) -> NormalizedInboundOrder:
        path = Path(path)
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[self.sheet_name] if self.sheet_name else wb[wb.sheetnames[0]]
        header_row, columns = self._find_header_row(ws)

        releases: List[NormalizedRelease] = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            po = clean_text(ws.cell(row_idx, columns["po"]).value)
            if not po:
                continue

            qty_raw = ws.cell(row_idx, columns["quantity"]).value
            try:
                source_quantity = float(qty_raw) if qty_raw not in (None, "") else None
            except (TypeError, ValueError):
                source_quantity = None

            source_uom = (
                clean_text(ws.cell(row_idx, columns.get("uom", 0)).value)
                if "uom" in columns
                else None
            )
            source_plant = (
                clean_text(ws.cell(row_idx, columns.get("plant", 0)).value)
                if "plant" in columns
                else None
            )

            pickup = excel_serial_to_datetime(
                ws.cell(row_idx, columns["pickup"]).value if "pickup" in columns else None
            )
            receive_dt = excel_serial_to_datetime(
                ws.cell(row_idx, columns["receive"]).value if "receive" in columns else None
            )

            review_reasons: List[str] = [
                "CanPack workbook is supplier/manufacturer-side schedule evidence; end-customer Sales Order context must be resolved",
                "Supplier quantity/UOM must not be sent to BC Sales Order until customer/item sales UOM mapping is proven",
            ]
            if source_quantity is None:
                review_reasons.append("Source quantity missing or non-numeric")
            if not clean_text(ws.cell(row_idx, columns["customer_item"]).value):
                review_reasons.append("Supplier customer material number missing")

            releases.append(
                NormalizedRelease(
                    # Retained for compatibility with the normalized release model. For this
                    # source format the value is a supplier-side PO/order reference and must not
                    # be assumed to be an end-customer PO until relationship resolution occurs.
                    customer_release_reference=po,
                    product_context=clean_text(ws.cell(row_idx, columns.get("description", 0)).value)
                    if "description" in columns else None,
                    customer_item_reference=clean_text(ws.cell(row_idx, columns["customer_item"]).value),
                    description=clean_text(ws.cell(row_idx, columns.get("description", 0)).value)
                    if "description" in columns else None,
                    # Do not promote CanPack's supplier-side call-off quantity/UOM into BC-ready
                    # Sales Order values. Preserve them as physical/source evidence only.
                    quantity=None,
                    uom=None,
                    physical_quantity=source_quantity,
                    physical_uom=source_uom,
                    source_facility_reference=source_plant,
                    requested_shipment_date=pickup,
                    requested_delivery_date=receive_dt.date() if receive_dt else None,
                    ship_to_candidate=None,
                    location_candidate=None,
                    source_coordinates=[f"{ws.title}!A{row_idx}:{ws.cell(row_idx, ws.max_column).coordinate}"],
                    quantity_source=f"{ws.title}!{ws.cell(row_idx, columns['quantity']).coordinate}",
                    quantity_resolution_method="supplier_source_requires_linked_customer_item_resolution",
                    extraction_confidence=1.0,
                    parser_review_reasons=review_reasons,
                )
            )

        return NormalizedInboundOrder(
            source=OrderSource(
                attachment_name=path.name,
                attachment_sha256=sha256_file(path),
                source_format=self.SOURCE_FORMAT,
                source_sheet=ws.title,
                source_party_role="SUPPLIER_MANUFACTURER",
            ),
            customer=OrderCustomer(
                candidate_customer_name=None,
                resolution_method="supplier_schedule_requires_end_customer_resolution",
                resolution_confidence=0.0,
                evidence=[
                    "Known CanPack workbook format is supplier/manufacturer-side schedule evidence",
                    "CanPack must not be assumed to be the BC sell-to customer",
                ],
            ),
            document=OrderDocument(document_type=DocumentType.SUPPLIER_SALES_ORDER_SCHEDULE),
            releases=releases,
            validation=OrderValidation(
                customer_status="UNRESOLVED_END_CUSTOMER",
                item_status="UNRESOLVED_CUSTOMER_ITEM_MAPPING",
                quantity_status="SOURCE_ONLY_REQUIRES_BC_SALES_RESOLUTION",
                ship_to_status="UNRESOLVED",
                location_status="UNRESOLVED",
                price_status="NOT_EVALUATED",
                exceptions=[
                    "Supplier schedule is not itself authorization to create a BC Sales Order",
                    "Resolve linked end customer, BC item, sales UOM, ship-to/location, and duplicate state first",
                ],
                proposed_action=ProposedAction.REVIEW,
            ),
        )
