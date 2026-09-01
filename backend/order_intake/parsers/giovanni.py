"""Deterministic parser for Giovanni monthly blanket OOR workbooks."""

from __future__ import annotations

import calendar
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from ..models import (
    DocumentType,
    NormalizedInboundOrder,
    NormalizedRelease,
    OrderCustomer,
    OrderDocument,
    OrderSource,
)
from ..utils import clean_text, excel_serial_to_datetime, numeric_reference, sha256_file


@dataclass(frozen=True)
class GiovanniQuantityProfile:
    product_context: str
    full_tl_quantity: Optional[float]
    customer_item_reference: Optional[str] = None
    uom: Optional[str] = None
    source: Optional[str] = None


class GiovanniOorParser:
    SOURCE_FORMAT = "GIOVANNI_OOR_XLSX"
    PRODUCTS = (
        "24oz Pasta",
        "24oz Salsa",
        "16oz Vinegar",
        "14oz Pizza",
        "16oz Salsa",
    )
    MONTHS = {name.lower(): idx for idx, name in enumerate(calendar.month_name) if name}

    def __init__(
        self,
        quantity_profiles: Optional[Dict[str, GiovanniQuantityProfile]] = None,
        sheet_name: str = "Orders",
    ):
        self.quantity_profiles = quantity_profiles or {}
        self.sheet_name = sheet_name

    @staticmethod
    def _period_parts(period: str) -> Tuple[int, int]:
        match = re.fullmatch(r"(\d{4})-(\d{2})", period.strip())
        if not match:
            raise ValueError("Giovanni period must be YYYY-MM")
        year, month = int(match.group(1)), int(match.group(2))
        if month < 1 or month > 12:
            raise ValueError("Giovanni period month must be 01-12")
        return year, month

    @staticmethod
    def _normalize(value) -> str:
        return " ".join(str(value or "").strip().split())

    def _match_section(self, value, target_month: int) -> Optional[str]:
        text = self._normalize(value)
        month_name = calendar.month_name[target_month]
        for product in self.PRODUCTS:
            if text.lower() == f"{month_name} {product}".lower():
                return product
        return None

    def _find_header_row(self, ws, section_row: int, start_col: int) -> Optional[int]:
        """Find the Load Count/Gamer PO/Gio PO header directly below a section."""
        for row_idx in range(section_row, min(section_row + 8, ws.max_row) + 1):
            values = [self._normalize(ws.cell(row_idx, start_col + offset).value).lower() for offset in range(5)]
            if (
                values[0] in {"load count", "load"}
                and values[1] == "gamer po"
                and values[2] == "gio po"
                and "delivery" in values[3]
            ):
                return row_idx
        return None

    @staticmethod
    def _semantic_note(bol_value, notes_value) -> Optional[str]:
        parts: List[str] = []
        bol_text = clean_text(bol_value)
        if bol_text:
            # Preserve free text entered in BOL #, but don't treat a plain numeric BOL as a note.
            stripped = bol_text.replace("-", "").replace(" ", "")
            if not stripped.isdigit():
                parts.append(bol_text)
        notes_text = clean_text(notes_value)
        if notes_text:
            parts.append(notes_text)
        return " | ".join(parts) or None

    def _quantity_for(self, product: str, note: Optional[str]) -> Tuple[Optional[float], Optional[str], Optional[str], List[str]]:
        profile = self.quantity_profiles.get(product)
        review: List[str] = []
        text = (note or "").lower()

        exception_tokens = (
            "pallet",
            "mixed",
            "32oz",
            "partial",
            "split",
            "cancel",
            "reroute",
        )
        if any(token in text for token in exception_tokens):
            review.append("Nonstandard/mixed/cancel/reroute note requires quantity/order review")
            return None, profile.customer_item_reference if profile else None, profile.uom if profile else None, review

        if not profile or profile.full_tl_quantity is None:
            review.append("No authoritative Giovanni full-TL quantity profile supplied")
            return None, profile.customer_item_reference if profile else None, profile.uom if profile else None, review

        return float(profile.full_tl_quantity), profile.customer_item_reference, profile.uom, review

    def parse(
        self,
        path: str | Path,
        *,
        period: str,
        include_existing: bool = False,
    ) -> NormalizedInboundOrder:
        """Parse releases for one target month/year.

        By default only candidate releases with a blank Gamer PO are returned. A
        BC duplicate check is still mandatory before any create operation.
        """
        path = Path(path)
        target_year, target_month = self._period_parts(period)
        wb = load_workbook(path, data_only=True, read_only=True)
        if self.sheet_name not in wb.sheetnames:
            raise ValueError(f"Giovanni workbook does not contain sheet '{self.sheet_name}'")
        ws = wb[self.sheet_name]

        sections: List[Tuple[int, int, str, int]] = []
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                product = self._match_section(ws.cell(row_idx, col_idx).value, target_month)
                if not product:
                    continue
                header_row = self._find_header_row(ws, row_idx, col_idx)
                if header_row:
                    sections.append((row_idx, col_idx, product, header_row))

        releases: List[NormalizedRelease] = []
        seen_coordinates = set()

        for section_row, start_col, product, header_row in sections:
            # Sections in the sample are compact. Stop at 100 rows or a new monthly
            # section in the same starting column, whichever comes first.
            for row_idx in range(header_row + 1, min(header_row + 100, ws.max_row) + 1):
                maybe_next_section = self._normalize(ws.cell(row_idx, start_col).value)
                if row_idx > header_row + 1 and any(
                    maybe_next_section.lower().startswith(f"{month.lower()} ")
                    for month in calendar.month_name[1:]
                ):
                    break

                load_raw = ws.cell(row_idx, start_col).value
                gamer_po = clean_text(ws.cell(row_idx, start_col + 1).value)
                gio_po_raw = ws.cell(row_idx, start_col + 2).value
                delivery_raw = ws.cell(row_idx, start_col + 3).value
                bol_raw = ws.cell(row_idx, start_col + 4).value
                notes_raw = ws.cell(row_idx, start_col + 5).value

                load_ref = numeric_reference(load_raw)
                gio_ref = numeric_reference(gio_po_raw)
                if not load_ref or not gio_ref:
                    continue  # skips SS and non-release rows

                delivery_dt = excel_serial_to_datetime(delivery_raw)
                if not delivery_dt or delivery_dt.year != target_year or delivery_dt.month != target_month:
                    continue

                if gamer_po and not include_existing:
                    continue

                coord = f"{ws.title}!{ws.cell(row_idx, start_col).coordinate}:{ws.cell(row_idx, start_col + 5).coordinate}"
                if coord in seen_coordinates:
                    continue
                seen_coordinates.add(coord)

                note = self._semantic_note(bol_raw, notes_raw)
                quantity, customer_item_ref, uom, review = self._quantity_for(product, note)

                note_lower = (note or "").lower()
                if "montreal" in note_lower:
                    review.append("Ship-to/location needs BC resolution")
                if product == "16oz Vinegar" and product not in self.quantity_profiles:
                    review.append("Confirm 16oz Vinegar quantity/item profile")

                profile = self.quantity_profiles.get(product)
                quantity_source = profile.source if profile else None

                releases.append(
                    NormalizedRelease(
                        customer_release_reference=gio_ref,
                        load_number=int(load_ref),
                        product_context=product,
                        customer_item_reference=customer_item_ref,
                        description=f"{calendar.month_name[target_month]} {product}",
                        quantity=quantity,
                        uom=uom,
                        requested_delivery_date=delivery_dt.date(),
                        existing_gamer_order=gamer_po,
                        notes=note,
                        source_coordinates=[coord],
                        quantity_source=quantity_source,
                        extraction_confidence=1.0,
                        parser_review_reasons=list(dict.fromkeys(review)),
                    )
                )

        return NormalizedInboundOrder(
            source=OrderSource(
                attachment_name=path.name,
                attachment_sha256=sha256_file(path),
                source_format=self.SOURCE_FORMAT,
                source_sheet=ws.title,
            ),
            customer=OrderCustomer(
                candidate_customer_name="Giovanni",
                resolution_method="known_format",
                resolution_confidence=1.0,
                evidence=[f"Known Giovanni monthly OOR format; period={period}"],
            ),
            document=OrderDocument(
                document_type=DocumentType.BLANKET_RELEASE_BATCH,
                period=period,
            ),
            releases=releases,
        )
