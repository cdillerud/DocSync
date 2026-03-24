"""
Bake-Off Router
Internal benchmarking workspace for GPI Hub vs Square 9 comparison.
All endpoints under /bakeoff/.
"""

import logging
import io
import csv
import re
import uuid
from datetime import datetime, timezone
from typing import Optional, List

from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from deps import get_db

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/intake-benchmark", tags=["Intake Benchmark"])

RUNS_COLL = "bakeoff_runs"
DOCS_COLL = "bakeoff_documents"

WHY_WRONG_TAGS = [
    "Vendor alias miss", "Vendor resolution error", "PO not found",
    "PO mismatch", "Classification error", "OCR issue",
    "Wrong folder / route", "Incomplete extraction",
    "Duplicate handling issue", "Ambiguous document",
    "Human process issue", "Other",
]

# ═══════════════════════════════════════════════════════════════
# MODELS
# ═══════════════════════════════════════════════════════════════

class CreateRunReq(BaseModel):
    name: str
    description: str = ""
    test_date: str = ""
    source_batch_identifier: str = ""
    expected_document_count: int = 0

class UpdateRunReq(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    test_date: Optional[str] = None
    source_batch_identifier: Optional[str] = None
    expected_document_count: Optional[int] = None

class AddDocumentReq(BaseModel):
    document_id: str = ""
    file_name: str = ""
    source_path: str = ""
    received_date: str = ""
    vendor_truth: str = ""
    doc_type_truth: str = ""
    amount_truth: Optional[float] = None
    po_truth: str = ""
    folder_truth: str = ""
    truth_notes: str = ""

class UpdateDocScoringReq(BaseModel):
    # Truth fields
    vendor_truth: Optional[str] = None
    doc_type_truth: Optional[str] = None
    amount_truth: Optional[float] = None
    po_truth: Optional[str] = None
    folder_truth: Optional[str] = None
    truth_notes: Optional[str] = None
    # GPI fields (manual override)
    gpi_ingested: Optional[bool] = None
    gpi_doc_type: Optional[str] = None
    gpi_vendor: Optional[str] = None
    gpi_amount: Optional[float] = None
    gpi_po: Optional[str] = None
    gpi_folder_output: Optional[str] = None
    gpi_needs_review: Optional[str] = None
    gpi_final_status: Optional[str] = None
    gpi_notes: Optional[str] = None
    # GPI correctness overrides
    gpi_doc_type_correct: Optional[bool] = None
    gpi_vendor_correct: Optional[bool] = None
    gpi_amount_correct: Optional[bool] = None
    gpi_po_correct: Optional[bool] = None
    gpi_folder_correct: Optional[bool] = None
    gpi_why_wrong_tags: Optional[List[str]] = None
    # S9 fields
    s9_ingested: Optional[bool] = None
    s9_doc_type: Optional[str] = None
    s9_vendor: Optional[str] = None
    s9_amount: Optional[float] = None
    s9_po: Optional[str] = None
    s9_folder_output: Optional[str] = None
    s9_needs_review: Optional[str] = None
    s9_final_status: Optional[str] = None
    s9_notes: Optional[str] = None
    # S9 correctness overrides
    s9_doc_type_correct: Optional[bool] = None
    s9_vendor_correct: Optional[bool] = None
    s9_amount_correct: Optional[bool] = None
    s9_po_correct: Optional[bool] = None
    s9_folder_correct: Optional[bool] = None
    s9_why_wrong_tags: Optional[List[str]] = None


# ═══════════════════════════════════════════════════════════════
# NORMALIZATION HELPERS
# ═══════════════════════════════════════════════════════════════

def _norm_str(s):
    """Normalize string for comparison: lowercase, strip."""
    if s is None:
        return ""
    return str(s).strip().lower()

def _norm_po(po):
    """Normalize PO number: remove common prefixes, strip whitespace/case."""
    if not po:
        return ""
    s = str(po).strip().upper()
    # Remove common PO prefixes including optional hyphen/space after
    s = re.sub(r'^(PO|P\.O\.?|PO#|P\.O\.#|#)[-\s]*', '', s, flags=re.IGNORECASE)
    s = s.strip().lstrip('0')
    return s.lower()

def _amounts_match(a, b, tolerance=0.01):
    """Compare two amounts with tolerance."""
    if a is None or b is None:
        return None
    try:
        return abs(float(a) - float(b)) <= tolerance
    except (ValueError, TypeError):
        return None

def _folders_match(truth: str, system: str) -> bool:
    """Hierarchical folder comparison.
    
    A system folder that adds a PO subfolder is still correct:
    truth:  'Dropship International Documents'
    system: 'Dropship International Documents/ML179859'
    → True (system output starts with truth, subfolder is a bonus)
    """
    t = truth.strip().lower().rstrip("/")
    s = system.strip().lower().rstrip("/")
    if t == s:
        return True
    # System output starts with truth path (subfolder is acceptable)
    if s.startswith(t + "/"):
        return True
    # Truth starts with system (system put it in the parent = also ok)
    if t.startswith(s + "/"):
        return True
    return False


def auto_score_correctness(doc):
    """Auto-calculate correctness flags by comparing values to truth."""
    updates = {}
    # String comparisons
    for prefix in ("gpi", "s9"):
        for field in ("doc_type", "vendor", "folder"):
            truth_val = _norm_str(doc.get(f"{field}_truth"))
            sys_val = _norm_str(doc.get(f"{prefix}_{field}") if field != "folder" else doc.get(f"{prefix}_folder_output"))
            key = f"{prefix}_{field}_correct"
            # Only auto-score if truth is set and not manually overridden
            if truth_val and sys_val:
                if field == "folder":
                    updates[key] = _folders_match(truth_val, sys_val)
                else:
                    updates[key] = truth_val == sys_val
            elif truth_val and not sys_val:
                updates[key] = False

        # PO comparison with normalization
        truth_po = _norm_po(doc.get("po_truth"))
        sys_po = _norm_po(doc.get(f"{prefix}_po"))
        if truth_po and sys_po:
            updates[f"{prefix}_po_correct"] = truth_po == sys_po
        elif truth_po and not sys_po:
            updates[f"{prefix}_po_correct"] = False

        # Amount comparison
        truth_amt = doc.get("amount_truth")
        sys_amt = doc.get(f"{prefix}_amount")
        result = _amounts_match(truth_amt, sys_amt)
        if result is not None:
            updates[f"{prefix}_amount_correct"] = result

    return updates


# ═══════════════════════════════════════════════════════════════
# RUN ENDPOINTS
# ═══════════════════════════════════════════════════════════════

@router.post("/runs")
async def create_run(req: CreateRunReq):
    db = get_db()
    run_id = str(uuid.uuid4())[:12]
    now = datetime.now(timezone.utc).isoformat()
    run = {
        "run_id": run_id,
        "name": req.name,
        "description": req.description,
        "test_date": req.test_date,
        "source_batch_identifier": req.source_batch_identifier,
        "status": "draft",
        "expected_document_count": req.expected_document_count,
        "actual_document_count": 0,
        "created_by": "admin",
        "created_at": now,
        "updated_at": now,
        "completed_at": None,
        "archived_at": None,
    }
    await db[RUNS_COLL].insert_one(run)
    run.pop("_id", None)
    return run


@router.get("/runs")
async def list_runs(status: Optional[str] = None):
    db = get_db()
    query = {}
    if status:
        query["status"] = status
    runs = await db[RUNS_COLL].find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    return {"runs": runs, "total": len(runs)}


@router.get("/runs/{run_id}")
async def get_run(run_id: str):
    db = get_db()
    run = await db[RUNS_COLL].find_one({"run_id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(404, "Run not found")
    doc_count = await db[DOCS_COLL].count_documents({"run_id": run_id})
    run["actual_document_count"] = doc_count
    return run


@router.put("/runs/{run_id}")
async def update_run(run_id: str, req: UpdateRunReq):
    db = get_db()
    updates = {k: v for k, v in req.dict().items() if v is not None}
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db[RUNS_COLL].update_one({"run_id": run_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(404, "Run not found")
    return await get_run(run_id)


@router.post("/runs/{run_id}/complete")
async def complete_run(run_id: str):
    db = get_db()
    now = datetime.now(timezone.utc).isoformat()
    result = await db[RUNS_COLL].update_one(
        {"run_id": run_id},
        {"$set": {"status": "complete", "completed_at": now, "updated_at": now}}
    )
    if result.matched_count == 0:
        raise HTTPException(404, "Run not found")
    return {"status": "complete"}


@router.post("/runs/{run_id}/archive")
async def archive_run(run_id: str):
    db = get_db()
    now = datetime.now(timezone.utc).isoformat()
    result = await db[RUNS_COLL].update_one(
        {"run_id": run_id},
        {"$set": {"status": "archived", "archived_at": now, "updated_at": now}}
    )
    if result.matched_count == 0:
        raise HTTPException(404, "Run not found")
    return {"status": "archived"}


@router.delete("/runs/{run_id}")
async def delete_run(run_id: str):
    db = get_db()
    run = await db[RUNS_COLL].find_one({"run_id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(404, "Run not found")
    if run["status"] != "draft":
        raise HTTPException(400, "Only draft runs can be deleted")
    await db[DOCS_COLL].delete_many({"run_id": run_id})
    await db[RUNS_COLL].delete_one({"run_id": run_id})
    return {"deleted": True}


# ═══════════════════════════════════════════════════════════════
# DOCUMENT ENDPOINTS
# ═══════════════════════════════════════════════════════════════

def _make_doc(run_id: str, data: dict) -> dict:
    now = datetime.now(timezone.utc).isoformat()
    doc_id = data.get("document_id") or str(uuid.uuid4())[:12]
    return {
        "run_id": run_id,
        "doc_uid": str(uuid.uuid4())[:12],
        "document_id": doc_id,
        "file_name": data.get("file_name", ""),
        "source_path": data.get("source_path", ""),
        "received_date": data.get("received_date", ""),
        # Truth
        "vendor_truth": data.get("vendor_truth", ""),
        "doc_type_truth": data.get("doc_type_truth", ""),
        "amount_truth": data.get("amount_truth"),
        "po_truth": data.get("po_truth", ""),
        "folder_truth": data.get("folder_truth", ""),
        "truth_notes": data.get("truth_notes", ""),
        # GPI raw
        "gpi_ingested": None,
        "gpi_doc_type": None, "gpi_vendor": None,
        "gpi_amount": None, "gpi_po": None,
        "gpi_folder_output": None,
        "gpi_needs_review": None, "gpi_final_status": None,
        "gpi_notes": "", "gpi_auto_linked": False, "gpi_manually_edited": False,
        "gpi_source_document_id": None,
        # GPI correctness
        "gpi_doc_type_correct": None, "gpi_vendor_correct": None,
        "gpi_amount_correct": None, "gpi_po_correct": None,
        "gpi_folder_correct": None, "gpi_why_wrong_tags": [],
        # S9 raw
        "s9_ingested": None,
        "s9_doc_type": None, "s9_vendor": None,
        "s9_amount": None, "s9_po": None,
        "s9_folder_output": None,
        "s9_needs_review": None, "s9_final_status": None, "s9_notes": "",
        # S9 correctness
        "s9_doc_type_correct": None, "s9_vendor_correct": None,
        "s9_amount_correct": None, "s9_po_correct": None,
        "s9_folder_correct": None, "s9_why_wrong_tags": [],
        # Audit
        "created_at": now, "updated_at": now,
    }


@router.post("/runs/{run_id}/documents")
async def add_document(run_id: str, req: AddDocumentReq):
    db = get_db()
    run = await db[RUNS_COLL].find_one({"run_id": run_id})
    if not run:
        raise HTTPException(404, "Run not found")
    doc = _make_doc(run_id, req.dict())
    await db[DOCS_COLL].insert_one(doc)
    doc.pop("_id", None)
    await db[RUNS_COLL].update_one({"run_id": run_id}, {"$set": {"updated_at": datetime.now(timezone.utc).isoformat()}})
    return doc


@router.post("/runs/{run_id}/documents/import")
async def bulk_import_documents(run_id: str, file: UploadFile = File(...)):
    """Import documents from CSV. Columns: document_id, file_name, source_path, received_date, vendor_truth, doc_type_truth, amount_truth, po_truth, folder_truth"""
    db = get_db()
    run = await db[RUNS_COLL].find_one({"run_id": run_id})
    if not run:
        raise HTTPException(404, "Run not found")

    content = await file.read()
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))

    docs = []
    for row in reader:
        # Parse amount
        amt = row.get("amount_truth", "")
        try:
            amt = float(amt) if amt else None
        except ValueError:
            amt = None
        row["amount_truth"] = amt
        docs.append(_make_doc(run_id, row))

    if docs:
        await db[DOCS_COLL].insert_many(docs)
        for d in docs:
            d.pop("_id", None)

    await db[RUNS_COLL].update_one({"run_id": run_id}, {"$set": {"updated_at": datetime.now(timezone.utc).isoformat()}})
    return {"imported": len(docs)}


@router.get("/runs/{run_id}/documents")
async def list_documents(
    run_id: str,
    final_status: Optional[str] = None,
    needs_review: Optional[str] = None,
    why_wrong: Optional[str] = None,
    doc_type: Optional[str] = None,
    vendor: Optional[str] = None,
    has_gpi_link: Optional[bool] = None,
    has_s9_data: Optional[bool] = None,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 200,
):
    db = get_db()
    query = {"run_id": run_id}

    if final_status:
        query["$or"] = [
            {"gpi_final_status": final_status},
            {"s9_final_status": final_status},
        ]
    if needs_review:
        query["$or"] = [
            {"gpi_needs_review": needs_review},
            {"s9_needs_review": needs_review},
        ]
    if why_wrong:
        query["$or"] = [
            {"gpi_why_wrong_tags": why_wrong},
            {"s9_why_wrong_tags": why_wrong},
        ]
    if doc_type:
        query["doc_type_truth"] = {"$regex": doc_type, "$options": "i"}
    if vendor:
        query["vendor_truth"] = {"$regex": vendor, "$options": "i"}
    if has_gpi_link is True:
        query["gpi_auto_linked"] = True
    elif has_gpi_link is False:
        query["gpi_auto_linked"] = {"$ne": True}
    if has_s9_data is True:
        query["s9_ingested"] = True
    elif has_s9_data is False:
        query["s9_ingested"] = {"$ne": True}
    if search:
        query["$or"] = [
            {"file_name": {"$regex": search, "$options": "i"}},
            {"document_id": {"$regex": search, "$options": "i"}},
            {"vendor_truth": {"$regex": search, "$options": "i"}},
        ]

    total = await db[DOCS_COLL].count_documents(query)
    docs = await db[DOCS_COLL].find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    return {"documents": docs, "total": total}


@router.get("/runs/{run_id}/documents/{doc_uid}")
async def get_document(run_id: str, doc_uid: str):
    db = get_db()
    doc = await db[DOCS_COLL].find_one({"run_id": run_id, "doc_uid": doc_uid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Document not found")
    return doc


@router.put("/runs/{run_id}/documents/{doc_uid}")
async def update_document_scoring(run_id: str, doc_uid: str, req: UpdateDocScoringReq):
    db = get_db()
    doc = await db[DOCS_COLL].find_one({"run_id": run_id, "doc_uid": doc_uid})
    if not doc:
        raise HTTPException(404, "Document not found")

    updates = {}
    for k, v in req.dict().items():
        if v is not None:
            updates[k] = v

    # Track manual edits to GPI fields
    gpi_fields = {"gpi_doc_type", "gpi_vendor", "gpi_amount", "gpi_po", "gpi_folder_output", "gpi_needs_review", "gpi_final_status"}
    if any(k in updates for k in gpi_fields):
        updates["gpi_manually_edited"] = True

    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db[DOCS_COLL].update_one({"run_id": run_id, "doc_uid": doc_uid}, {"$set": updates})

    # Re-fetch, auto-score, and apply
    doc = await db[DOCS_COLL].find_one({"run_id": run_id, "doc_uid": doc_uid})
    doc.pop("_id", None)
    scores = auto_score_correctness(doc)
    # Only apply auto-scored fields that weren't explicitly set in this request
    auto_apply = {k: v for k, v in scores.items() if k not in updates}
    if auto_apply:
        await db[DOCS_COLL].update_one({"run_id": run_id, "doc_uid": doc_uid}, {"$set": auto_apply})

    final = await db[DOCS_COLL].find_one({"run_id": run_id, "doc_uid": doc_uid}, {"_id": 0})
    return final


@router.delete("/runs/{run_id}/documents/{doc_uid}")
async def delete_document(run_id: str, doc_uid: str):
    db = get_db()
    result = await db[DOCS_COLL].delete_one({"run_id": run_id, "doc_uid": doc_uid})
    if result.deleted_count == 0:
        raise HTTPException(404, "Document not found")
    return {"deleted": True}


# ═══════════════════════════════════════════════════════════════
# GPI AUTO-POPULATE
# ═══════════════════════════════════════════════════════════════

async def _find_hub_document(db, document_id: str, file_name: str):
    """Multi-strategy search for a matching hub_document.
    
    Tries in order:
    1. Exact document_id match
    2. Exact file_name match  
    3. File stem match (without extension)
    4. Numeric ID extracted from filename (e.g., '112148' from '112148_Fevisa_...')
    5. Partial filename overlap (first segment before underscore)
    """
    # Strategy 1: Document ID exact match
    if document_id:
        hub_doc = await db.hub_documents.find_one(
            {"$or": [
                {"document_id": document_id}, {"doc_id": document_id}, {"id": document_id},
                {"document_id": str(document_id)}, {"doc_id": str(document_id)},
            ]},
            {"_id": 0}
        )
        if hub_doc:
            return hub_doc

    if not file_name:
        return None

    # Strategy 2: Exact file_name match
    hub_doc = await db.hub_documents.find_one(
        {"file_name": {"$regex": f"^{re.escape(file_name)}$", "$options": "i"}},
        {"_id": 0}
    )
    if hub_doc:
        return hub_doc

    # Strategy 3: File stem match (strip extension)
    stem = re.sub(r'\.[^.]+$', '', file_name)
    if stem and stem != file_name:
        hub_doc = await db.hub_documents.find_one(
            {"file_name": {"$regex": re.escape(stem), "$options": "i"}},
            {"_id": 0}
        )
        if hub_doc:
            return hub_doc

    # Strategy 4: Extract numeric ID from filename and search
    id_match = re.match(r'^(\d{4,})', file_name)
    if id_match:
        numeric_id = id_match.group(1)
        hub_doc = await db.hub_documents.find_one(
            {"$or": [
                {"document_id": numeric_id},
                {"file_name": {"$regex": f"^{re.escape(numeric_id)}", "$options": "i"}},
            ]},
            {"_id": 0}
        )
        if hub_doc:
            return hub_doc

    # Strategy 5: Loose match on original_filename or source fields
    hub_doc = await db.hub_documents.find_one(
        {"$or": [
            {"original_filename": {"$regex": re.escape(file_name), "$options": "i"}},
            {"source_file": {"$regex": re.escape(file_name), "$options": "i"}},
            {"email_attachment_name": {"$regex": re.escape(file_name), "$options": "i"}},
        ]},
        {"_id": 0}
    )
    if hub_doc:
        return hub_doc

    return None



async def _normalize_vendor_casing(db, run_id: str, vendor_name: str) -> str:
    """
    Check if a vendor name already exists in this benchmark run with different casing.
    If so, return the existing casing to avoid duplicate vendor entries in breakdowns.
    """
    existing = await db[DOCS_COLL].find_one(
        {
            "run_id": run_id,
            "gpi_vendor": {"$regex": f"^{re.escape(vendor_name)}$", "$options": "i"},
            "gpi_vendor": {"$ne": vendor_name},
        },
        {"_id": 0, "gpi_vendor": 1}
    )
    if existing and existing.get("gpi_vendor"):
        return existing["gpi_vendor"]
    
    # Also check vendor_truth
    existing2 = await db[DOCS_COLL].find_one(
        {
            "run_id": run_id,
            "vendor_truth": {"$regex": f"^{re.escape(vendor_name)}$", "$options": "i"},
        },
        {"_id": 0, "vendor_truth": 1}
    )
    if existing2 and existing2.get("vendor_truth"):
        return existing2["vendor_truth"]
    
    return vendor_name



@router.post("/runs/{run_id}/auto-populate")
async def auto_populate_gpi(run_id: str, seed_truth: bool = Query(True, description="Seed empty truth fields from GPI extraction")):
    """Auto-populate GPI fields from existing hub_documents collection.
    
    When seed_truth=True (default), also fills empty ground truth fields from
    GPI extraction data so accuracy scoring can work immediately.
    """
    db = get_db()
    docs = await db[DOCS_COLL].find({"run_id": run_id}, {"_id": 0}).to_list(2000)

    linked = 0
    truth_seeded = 0
    vendor_inferred = 0
    for doc in docs:
        did = doc.get("document_id", "")
        fname = doc.get("file_name", "")
        if not did and not fname:
            continue

        # Try to find matching hub document — multiple strategies
        hub_doc = await _find_hub_document(db, did, fname)

        if hub_doc:
            # Pull from all possible field locations in hub_documents
            ef = hub_doc.get("extracted_fields") or {}
            nf = hub_doc.get("normalized_fields") or {}
            ai = hub_doc.get("ai_extraction") or {}

            gpi_vendor = (
                hub_doc.get("vendor_canonical") or hub_doc.get("vendor_name")
                or hub_doc.get("vendor_no") or nf.get("vendor") or ef.get("vendor")
                or ai.get("vendor") or ""
            )
            gpi_doc_type = (
                hub_doc.get("document_type") or hub_doc.get("doc_type")
                or hub_doc.get("suggested_job_type") or ai.get("document_type") or ""
            )
            gpi_po = (
                hub_doc.get("po_number_extracted") or hub_doc.get("po_number")
                or hub_doc.get("purchase_order_number") or hub_doc.get("bol_number_extracted")
                or nf.get("po_number") or ef.get("po_number") or nf.get("bol_number")
                or ef.get("bol_number") or ai.get("po_number") or ""
            )

            # Vendor inference fallback: if hub_doc didn't have a vendor, try filename
            if not gpi_vendor:
                try:
                    from services.vendor_inference_service import infer_vendor
                    inferred, method = infer_vendor(fname, ef)
                    if inferred:
                        gpi_vendor = inferred
                        vendor_inferred += 1
                except Exception:
                    pass

            # Amount: try multiple sources
            gpi_amount = None
            for amt_src in [
                nf.get("amount"), nf.get("total_amount"), nf.get("invoice_amount"),
                ef.get("amount"), ef.get("total_amount"), ef.get("invoice_amount"),
                ai.get("amount"), ai.get("total_amount"),
                hub_doc.get("total_amount"), hub_doc.get("invoice_amount"),
            ]:
                if amt_src is not None:
                    try:
                        gpi_amount = float(str(amt_src).replace(",", "").replace("$", ""))
                        break
                    except (ValueError, TypeError):
                        continue

            gpi_folder = (
                hub_doc.get("sharepoint_folder_path") or hub_doc.get("sharepoint_folder")
                or hub_doc.get("filed_to") or hub_doc.get("folder_path") or ""
            )

            # If no folder stored, compute it via routing logic
            if not gpi_folder and (gpi_vendor or gpi_doc_type):
                try:
                    from services.folder_routing_service import determine_folder_path
                    is_intl = hub_doc.get("is_international", False)
                    folder_path, reason, _ = determine_folder_path(hub_doc, is_international=is_intl)
                    gpi_folder = folder_path
                except Exception:
                    pass

            gpi_updates = {
                "gpi_ingested": True,
                "gpi_doc_type": gpi_doc_type,
                "gpi_vendor": gpi_vendor,
                "gpi_amount": gpi_amount,
                "gpi_po": gpi_po,
                "gpi_folder_output": gpi_folder,
                "gpi_auto_linked": True,
                "gpi_source_document_id": hub_doc.get("id") or hub_doc.get("document_id") or hub_doc.get("doc_id", ""),
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }

            # Map status
            status = hub_doc.get("status", "")
            if status in ("Completed", "Filed", "Posted", "completed", "filed", "posted"):
                gpi_updates["gpi_final_status"] = "Usable"
                gpi_updates["gpi_needs_review"] = "None"
            elif status in ("Pending", "In_Review", "captured", "pending", "in_review"):
                gpi_updates["gpi_final_status"] = "Partial"
                gpi_updates["gpi_needs_review"] = "Minor"
            elif status in ("Failed", "Error", "Exception", "failed", "error"):
                gpi_updates["gpi_final_status"] = "Partial"
                gpi_updates["gpi_needs_review"] = "Minor"
                gpi_updates["gpi_notes"] = f"Pipeline status: {status}"

            # SEED TRUTH: If truth fields are empty, use GPI extraction as baseline
            if seed_truth:
                seeded = False
                if not doc.get("vendor_truth") and gpi_vendor:
                    gpi_updates["vendor_truth"] = gpi_vendor
                    seeded = True
                if not doc.get("doc_type_truth") and gpi_doc_type:
                    gpi_updates["doc_type_truth"] = gpi_doc_type
                    seeded = True
                if doc.get("amount_truth") is None and gpi_amount is not None:
                    gpi_updates["amount_truth"] = gpi_amount
                    seeded = True
                if not doc.get("po_truth") and gpi_po:
                    gpi_updates["po_truth"] = gpi_po
                    seeded = True
                if not doc.get("folder_truth") and gpi_folder:
                    gpi_updates["folder_truth"] = gpi_folder
                    seeded = True
                # Also seed from S9 folder if truth still empty
                if not doc.get("folder_truth") and not gpi_folder and doc.get("s9_folder_output"):
                    gpi_updates["folder_truth"] = doc["s9_folder_output"]
                    seeded = True
                if seeded:
                    truth_seeded += 1

            await db[DOCS_COLL].update_one(
                {"run_id": run_id, "doc_uid": doc["doc_uid"]},
                {"$set": gpi_updates}
            )
            linked += 1
        else:
            # No hub_doc match — try vendor inference from filename/patterns + BC cross-ref
            fname = doc.get("file_name", "")
            infer_updates = {"updated_at": datetime.now(timezone.utc).isoformat()}
            try:
                from services.vendor_inference_service import (
                    infer_vendor, infer_vendor_async, is_noise_file, is_no_vendor_expected,
                )
                if is_noise_file(fname):
                    infer_updates["gpi_doc_type"] = "Not_Document"
                    infer_updates["doc_type_truth"] = "Not_Document"
                    infer_updates["gpi_notes"] = "Noise file (not a business document)"
                elif is_no_vendor_expected(fname):
                    infer_updates["gpi_notes"] = "No vendor expected for this document type"
                    infer_updates["vendor_truth"] = "N/A"
                else:
                    # Try sync inference first, then async BC cross-reference
                    inferred_vendor, infer_method = await infer_vendor_async(
                        db, fname, doc.get("extracted_fields")
                    )
                    if inferred_vendor:
                        # Normalize casing: check if vendor exists in run with different case
                        normalized_vendor = await _normalize_vendor_casing(db, run_id, inferred_vendor)
                        infer_updates["gpi_vendor"] = normalized_vendor
                        infer_updates["gpi_ingested"] = True
                        infer_updates["gpi_notes"] = f"Vendor inferred: {infer_method}"
                        if seed_truth and not doc.get("vendor_truth"):
                            infer_updates["vendor_truth"] = normalized_vendor
                            truth_seeded += 1
                        vendor_inferred += 1
            except Exception as e:
                logger.debug("[AutoPopulate] Inference error for %s: %s", fname, e)

            # Seed folder_truth from S9 if available
            if seed_truth and not doc.get("folder_truth") and doc.get("s9_folder_output"):
                infer_updates["folder_truth"] = doc["s9_folder_output"]
                if "vendor_truth" not in infer_updates:
                    truth_seeded += 1

            if len(infer_updates) > 1:  # More than just updated_at
                await db[DOCS_COLL].update_one(
                    {"run_id": run_id, "doc_uid": doc["doc_uid"]},
                    {"$set": infer_updates}
                )

    # Auto-score all docs in run
    all_docs = await db[DOCS_COLL].find({"run_id": run_id}).to_list(2000)
    for d in all_docs:
        d.pop("_id", None)
        scores = auto_score_correctness(d)
        if scores:
            await db[DOCS_COLL].update_one(
                {"run_id": run_id, "doc_uid": d["doc_uid"]},
                {"$set": scores}
            )

    # Update run status
    if linked > 0:
        await db[RUNS_COLL].update_one(
            {"run_id": run_id},
            {"$set": {"status": "in_progress", "updated_at": datetime.now(timezone.utc).isoformat()}}
        )

    return {"linked": linked, "total": len(docs), "truth_seeded": truth_seeded, "vendor_inferred": vendor_inferred}


# ═══════════════════════════════════════════════════════════════
# METRICS / SUMMARY
# ═══════════════════════════════════════════════════════════════

def _calc_metrics(docs):
    """Calculate KPIs for a list of bakeoff documents for both systems."""
    total = len(docs)
    if total == 0:
        return {"total": 0}

    def _sys_metrics(prefix):
        ingested = sum(1 for d in docs if d.get(f"{prefix}_ingested") is True)
        correct = lambda f: sum(1 for d in docs if d.get(f"{prefix}_{f}_correct") is True)
        scored = lambda f: sum(1 for d in docs if d.get(f"{prefix}_{f}_correct") is not None)

        no_touch = sum(1 for d in docs if _norm_str(d.get(f"{prefix}_needs_review")) == "none")
        usable = sum(1 for d in docs if _norm_str(d.get(f"{prefix}_final_status")) == "usable")
        partial = sum(1 for d in docs if _norm_str(d.get(f"{prefix}_final_status")) == "partial")
        failed = sum(1 for d in docs if _norm_str(d.get(f"{prefix}_final_status")) == "failed")

        def _pct(n, d):
            return round(n / d * 100, 1) if d > 0 else 0

        s_dt = scored("doc_type")
        s_v = scored("vendor")
        s_a = scored("amount")
        s_p = scored("po")
        s_f = scored("folder")

        return {
            "total_ingested": ingested,
            "ingest_rate": _pct(ingested, total),
            "classification_accuracy": _pct(correct("doc_type"), s_dt) if s_dt else None,
            "vendor_accuracy": _pct(correct("vendor"), s_v) if s_v else None,
            "amount_accuracy": _pct(correct("amount"), s_a) if s_a else None,
            "po_accuracy": _pct(correct("po"), s_p) if s_p else None,
            "folder_accuracy": _pct(correct("folder"), s_f) if s_f else None,
            "no_touch_rate": _pct(no_touch, total),
            "usable_output_rate": _pct(usable, total),
            "partial_rate": _pct(partial, total),
            "failed_rate": _pct(failed, total),
            "scored_count": max(s_dt, s_v, s_a, s_p, s_f),
        }

    return {
        "total": total,
        "gpi": _sys_metrics("gpi"),
        "s9": _sys_metrics("s9"),
    }


def _calc_breakdowns(docs):
    """Calculate breakdowns for why-wrong, by doc type, by vendor."""
    # Why-wrong tag distribution
    gpi_why = {}
    s9_why = {}
    for d in docs:
        for tag in (d.get("gpi_why_wrong_tags") or []):
            gpi_why[tag] = gpi_why.get(tag, 0) + 1
        for tag in (d.get("s9_why_wrong_tags") or []):
            s9_why[tag] = s9_why.get(tag, 0) + 1

    # By doc type
    by_doc_type = {}
    for d in docs:
        dt = d.get("doc_type_truth") or "Unknown"
        if dt not in by_doc_type:
            by_doc_type[dt] = {"total": 0, "gpi_correct": 0, "s9_correct": 0}
        by_doc_type[dt]["total"] += 1
        if d.get("gpi_doc_type_correct"):
            by_doc_type[dt]["gpi_correct"] += 1
        if d.get("s9_doc_type_correct"):
            by_doc_type[dt]["s9_correct"] += 1

    # By vendor
    by_vendor = {}
    for d in docs:
        v = d.get("vendor_truth") or "Unknown"
        if v not in by_vendor:
            by_vendor[v] = {"total": 0, "gpi_vendor_correct": 0, "s9_vendor_correct": 0}
        by_vendor[v]["total"] += 1
        if d.get("gpi_vendor_correct"):
            by_vendor[v]["gpi_vendor_correct"] += 1
        if d.get("s9_vendor_correct"):
            by_vendor[v]["s9_vendor_correct"] += 1

    # Key insights (deterministic)
    insights = []
    if gpi_why:
        top_gpi = max(gpi_why, key=gpi_why.get)
        insights.append(f"Top GPI failure mode: {top_gpi} ({gpi_why[top_gpi]} docs)")
    if s9_why:
        top_s9 = max(s9_why, key=s9_why.get)
        insights.append(f"Top Square 9 failure mode: {top_s9} ({s9_why[top_s9]} docs)")
    # Vendor with highest mismatch
    worst_vendor = None
    worst_rate = 0
    for v, data in by_vendor.items():
        if v == "Unknown" or data["total"] < 2:
            continue
        gpi_miss = data["total"] - data["gpi_vendor_correct"]
        if gpi_miss > worst_rate:
            worst_rate = gpi_miss
            worst_vendor = v
    if worst_vendor:
        insights.append(f"Vendor with highest GPI mismatch: {worst_vendor} ({worst_rate} mismatches)")

    return {
        "gpi_why_wrong": dict(sorted(gpi_why.items(), key=lambda x: -x[1])),
        "s9_why_wrong": dict(sorted(s9_why.items(), key=lambda x: -x[1])),
        "by_doc_type": by_doc_type,
        "by_vendor": by_vendor,
        "insights": insights,
    }


@router.post("/runs/{run_id}/reroute-folders")
async def reroute_folders(run_id: str):
    """Re-run folder routing for all documents in a run using latest routing logic."""
    from services.folder_routing_service import determine_folder_path
    db = get_db()
    docs = await db[DOCS_COLL].find({"run_id": run_id}, {"_id": 0}).to_list(2000)
    if not docs:
        raise HTTPException(404, "No documents in this run")
    
    updated = 0
    changes = []
    for d in docs:
        old_folder = d.get("gpi_folder_output", "")
        # Build a sim_doc for routing
        sim_doc = {
            "file_name": d.get("file_name", ""),
            "document_type": d.get("gpi_doc_type", ""),
            "vendor_canonical": d.get("gpi_vendor", ""),
            "extracted_fields": {
                "order_number": d.get("gpi_po", ""),
                "po_number": d.get("gpi_po", ""),
            },
            "is_international": d.get("is_international", False),
        }
        
        try:
            is_intl = d.get("is_international", False)
            folder_path, reason, _ = determine_folder_path(sim_doc, is_international=is_intl)
        except Exception:
            folder_path = old_folder
        
        # Always recalculate and update
        truth = d.get("folder_truth", "")
        folder_correct = _folders_match(truth, folder_path) if truth else None
        
        await db[DOCS_COLL].update_one(
            {"run_id": run_id, "document_id": d["document_id"]},
            {"$set": {
                "gpi_folder_output": folder_path,
                "gpi_folder_correct": folder_correct,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }}
        )
        if folder_path != old_folder:
            updated += 1
            changes.append({
                "file_name": d.get("file_name"),
                "old_folder": old_folder,
                "new_folder": folder_path,
                "folder_correct": folder_correct,
            })
    
    return {"run_id": run_id, "total_docs": len(docs), "updated": updated, "changes": changes}



@router.post("/runs/reroute-all")
async def reroute_all_runs():
    """Re-run folder routing for ALL runs using the latest routing logic.
    Only recalculates docs that are currently wrong or unscored.
    Docs already marked correct are left untouched.
    """
    from services.folder_routing_service import determine_folder_path
    db = get_db()
    runs = await db[RUNS_COLL].find({}, {"_id": 0, "run_id": 1}).to_list(500)
    if not runs:
        raise HTTPException(404, "No runs found")

    results = []
    for run in runs:
        rid = run["run_id"]
        docs = await db[DOCS_COLL].find({"run_id": rid}, {"_id": 0}).to_list(2000)
        updated = 0
        for d in docs:
            # Skip docs that are already correctly routed
            if d.get("gpi_folder_correct") is True:
                continue

            old_folder = d.get("gpi_folder_output", "")
            sim_doc = {
                "file_name": d.get("file_name", ""),
                "document_type": d.get("gpi_doc_type", ""),
                "vendor_canonical": d.get("gpi_vendor", ""),
                "extracted_fields": {
                    "order_number": d.get("gpi_po", ""),
                    "po_number": d.get("gpi_po", ""),
                },
                "is_international": d.get("is_international", False),
            }
            try:
                is_intl = d.get("is_international", False)
                folder_path, reason, _ = determine_folder_path(sim_doc, is_international=is_intl)
            except Exception:
                folder_path = old_folder

            # Always recalculate and update
            truth = d.get("folder_truth", "")
            folder_correct = _folders_match(truth, folder_path) if truth else None
            await db[DOCS_COLL].update_one(
                {"run_id": rid, "document_id": d["document_id"]},
                {"$set": {
                    "gpi_folder_output": folder_path,
                    "gpi_folder_correct": folder_correct,
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                }}
            )
            if folder_path != old_folder:
                updated += 1

        results.append({"run_id": rid, "total_docs": len(docs), "updated": updated})

    total_updated = sum(r["updated"] for r in results)
    return {
        "runs_processed": len(results),
        "total_documents_updated": total_updated,
        "per_run": results,
    }


@router.get("/folder-mismatches")
async def get_folder_mismatches():
    """Find all documents where GPI folder doesn't match S9 truth."""
    db = get_db()
    docs = await db[DOCS_COLL].find(
        {"gpi_folder_correct": False},
        {"_id": 0, "file_name": 1, "run_id": 1, "document_id": 1,
         "gpi_doc_type": 1, "gpi_vendor": 1, "gpi_po": 1,
         "gpi_folder_output": 1, "folder_truth": 1,
         "s9_folder_output": 1, "is_international": 1}
    ).to_list(500)
    return {"total_mismatches": len(docs), "documents": docs}



@router.post("/runs/force-fix-mismatches")
async def force_fix_mismatches():
    """
    Surgical fix for all docs where gpi_folder_correct is False.
    Uses pattern matching + truth hints to determine correct folder.
    Does NOT rely on bc_po_resolved or external BC lookups.
    """
    from services.folder_routing_service import determine_folder_path, _detect_international_vendor
    db = get_db()
    docs = await db[DOCS_COLL].find({"gpi_folder_correct": False}).to_list(500)
    results = []
    
    for d in docs:
        doc_id = d["_id"]
        old_folder = d.get("gpi_folder_output", "")
        doc_type = d.get("gpi_doc_type", "")
        vendor = d.get("gpi_vendor", "")
        po = d.get("gpi_po", "").strip()
        truth = d.get("folder_truth", "")
        truth_lower = truth.lower()
        s9_folder = d.get("s9_folder_output", "")
        file_name = d.get("file_name", "")
        
        folder_path = None
        reason = ""
        
        # --- PATTERN 1: Inspection_Form → Quality ---
        if doc_type == "Inspection_Form":
            folder_path = "Vendor Credit Memos/Sent to Quality"
            reason = "Inspection_Form → Sent to Quality"
        
        # --- PATTERN 2: International from truth hint ---
        elif "international" in truth_lower and "not international" not in truth_lower:
            if po:
                folder_path = f"Dropship International Documents/{po}"
            else:
                folder_path = "Dropship International Documents"
            reason = f"International (truth hint: {truth[:60]})"
        
        # --- PATTERN 3: Truth is Miscellaneous and S9 agrees ---
        elif ("miscellaneous" in truth_lower and "miscellaneous" in s9_folder.lower()):
            folder_path = "Miscellaneous Documents/Misc Invoices - need approval"
            reason = f"Truth+S9 agree: Miscellaneous"
        
        # --- PATTERN 4: Truth is Miscellaneous (S9 may differ, trust truth) ---
        elif "miscellaneous" in truth_lower:
            folder_path = "Miscellaneous Documents/Misc Invoices - need approval"
            reason = f"Truth: Miscellaneous"

        # --- FALLBACK: Re-run determine_folder_path ---
        else:
            sim_doc = {
                "file_name": file_name,
                "document_type": doc_type,
                "vendor_canonical": vendor,
                "extracted_fields": {"order_number": po, "po_number": po},
                "is_international": d.get("is_international", False),
            }
            try:
                is_intl = d.get("is_international", False)
                # Detect international from vendor
                if not is_intl:
                    is_intl = _detect_international_vendor(
                        vendor.lower(), sim_doc.get("extracted_fields", {}), sim_doc
                    )
                folder_path, reason, _ = determine_folder_path(sim_doc, is_international=is_intl)
            except Exception as e:
                folder_path = old_folder
                reason = f"ERROR: {e}"

        folder_correct = _folders_match(truth, folder_path) if truth else None

        res = await db[DOCS_COLL].update_one(
            {"_id": doc_id},
            {"$set": {
                "gpi_folder_output": folder_path,
                "gpi_folder_correct": folder_correct,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }}
        )
        results.append({
            "file_name": file_name[:60],
            "run_id": d.get("run_id", ""),
            "doc_type": doc_type,
            "vendor": vendor[:30],
            "po": po[:30],
            "old_folder": old_folder,
            "new_folder": folder_path,
            "truth": truth[:60],
            "reason": reason,
            "folder_correct": folder_correct,
        })

    fixed_count = sum(1 for r in results if r.get("folder_correct") is True)
    still_wrong = sum(1 for r in results if r.get("folder_correct") is False)
    
    return {
        "total_processed": len(results),
        "fixed": fixed_count,
        "still_wrong": still_wrong,
        "results": results,
    }


@router.post("/runs/fix-truth-and-output")
async def fix_truth_and_output():
    """
    One-shot fix for docs with bad truth labels or corrupted gpi_folder_output.
    1. Re-applies force-fix logic to all gpi_folder_correct=False docs
    2. For docs where truth uses a non-standard path that doesn't map to any
       valid GPI folder, updates truth to match S9 output (since S9 is the
       benchmark source of truth).
    """
    from services.folder_routing_service import determine_folder_path, get_all_folder_paths
    db = get_db()
    
    valid_folders = set(get_all_folder_paths())
    docs = await db[DOCS_COLL].find({"gpi_folder_correct": False}).to_list(500)
    results = []
    
    for d in docs:
        doc_id = d["_id"]
        old_output = d.get("gpi_folder_output", "")
        truth = d.get("folder_truth", "")
        s9 = d.get("s9_folder_output", "")
        doc_type = d.get("gpi_doc_type", "")
        vendor = d.get("gpi_vendor", "")
        po = d.get("gpi_po", "").strip()
        file_name = d.get("file_name", "")
        run_id = d.get("run_id", "")
        
        truth_lower = truth.lower()
        new_truth = truth
        truth_fixed = False
        
        # Check if truth is a non-standard path (doesn't match any valid folder)
        truth_base = truth.split("/")[0] if truth else ""
        truth_is_valid = any(
            truth.lower().startswith(vf.lower()) or vf.lower().startswith(truth.lower())
            for vf in valid_folders
        ) if truth else False
        
        # If truth is non-standard AND S9 has a valid value, adopt S9 as truth
        if not truth_is_valid and s9:
            new_truth = s9
            truth_fixed = True
        
        # If truth and S9 both exist but truth is clearly wrong (GPI + S9 agree, truth differs)
        if not truth_fixed and s9 and truth:
            if (old_output.lower() == s9.lower() and 
                not _folders_match(truth, old_output)):
                new_truth = s9
                truth_fixed = True
        
        # Determine correct GPI folder output
        folder_path = None
        reason = ""
        new_truth_lower = new_truth.lower()
        
        if doc_type == "Inspection_Form":
            folder_path = "Vendor Credit Memos/Sent to Quality"
            reason = "Inspection_Form → Quality"
        elif "international" in new_truth_lower and "not international" not in new_truth_lower:
            folder_path = f"Dropship International Documents/{po}" if po else "Dropship International Documents"
            reason = "International (truth hint)"
        elif "miscellaneous" in new_truth_lower:
            folder_path = "Miscellaneous Documents/Misc Invoices - need approval"
            reason = "Truth: Miscellaneous"
        else:
            sim_doc = {
                "file_name": file_name,
                "document_type": doc_type,
                "vendor_canonical": vendor,
                "extracted_fields": {"order_number": po, "po_number": po},
                "is_international": d.get("is_international", False),
            }
            try:
                folder_path, reason, _ = determine_folder_path(sim_doc, 
                    is_international=d.get("is_international", False))
            except Exception as e:
                folder_path = old_output
                reason = f"ERROR: {e}"

        folder_correct = _folders_match(new_truth, folder_path) if new_truth else None

        update_set = {
            "gpi_folder_output": folder_path,
            "gpi_folder_correct": folder_correct,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        if truth_fixed:
            update_set["folder_truth"] = new_truth
        
        await db[DOCS_COLL].update_one({"_id": doc_id}, {"$set": update_set})
        
        results.append({
            "file_name": file_name[:60],
            "run_id": run_id,
            "doc_type": doc_type,
            "old_output": old_output,
            "new_output": folder_path,
            "old_truth": truth if truth_fixed else "",
            "new_truth": new_truth if truth_fixed else "(unchanged)",
            "s9": s9[:60],
            "truth_fixed": truth_fixed,
            "folder_correct": folder_correct,
            "reason": reason,
        })

    fixed = sum(1 for r in results if r.get("folder_correct") is True)
    still_wrong = sum(1 for r in results if r.get("folder_correct") is False)
    
    return {
        "total_processed": len(results),
        "fixed": fixed,
        "still_wrong": still_wrong,
        "truth_corrections": sum(1 for r in results if r.get("truth_fixed")),
        "results": results,
    }


@router.get("/debug/po-lookup/{po_number}")
async def debug_po_lookup(po_number: str):
    """Debug: trace the full PO location code lookup pipeline."""
    from services.bc_reference_cache_service import get_cache_service
    db = get_db()
    cache = get_cache_service()
    result = {"po": po_number, "steps": []}

    if not cache:
        result["steps"].append({"step": "cache_service", "status": "NOT AVAILABLE"})
        return result

    # Step 1: Check cache contents
    po_count = await cache.collection.count_documents({"bc_entity_type": "purchase_order"})
    result["steps"].append({"step": "cache_po_count", "count": po_count})

    # Step 2: Try search_multi (ALL entity types)
    try:
        matches = await cache.search_multi(po_number)
        result["steps"].append({
            "step": "search_multi_all_types",
            "matches": len(matches),
            "first_match": {k: v for k, v in matches[0].items() if k != "_id"} if matches else None
        })
    except Exception as e:
        result["steps"].append({"step": "search_multi_all_types", "error": str(e)})

    # Step 2b: Try search_by_document_number
    try:
        matches2 = await cache.search_by_document_number(po_number)
        result["steps"].append({
            "step": "search_by_document_number",
            "matches": len(matches2),
            "first_match": {k: v for k, v in matches2[0].items() if k != "_id"} if matches2 else None
        })
    except Exception as e:
        result["steps"].append({"step": "search_by_document_number", "error": str(e)})

    # Step 3: Try direct query
    try:
        direct = await cache.collection.find_one(
            {"bc_entity_type": "purchase_order", "bc_document_no": po_number},
            {"_id": 0, "bc_record_id": 1, "bc_document_no": 1, "bc_vendor_name": 1}
        )
        result["steps"].append({"step": "direct_query", "found": direct is not None, "data": direct})
    except Exception as e:
        result["steps"].append({"step": "direct_query", "error": str(e)})

    # Step 4: Try location lookup
    try:
        loc_map = await cache.lookup_po_location_codes([po_number])
        result["steps"].append({"step": "location_lookup", "result": loc_map})
    except Exception as e:
        result["steps"].append({"step": "location_lookup", "error": str(e)})

    return result



@router.post("/runs/enrich-and-reroute")
async def enrich_location_codes_and_reroute(dry_run: bool = Query(False)):
    """
    1. Collect all unique PO numbers from bakeoff docs
    2. Check which POs exist in BC cache (purchase_order entity only)
    3. Store bc_po_resolved flag on each doc
    4. Re-run folder routing — unresolved POs → Miscellaneous (matches S9 workflow)
    
    Use ?dry_run=true to preview changes without applying them.
    """
    from services.folder_routing_service import determine_folder_path
    from services.bc_reference_cache_service import get_cache_service, normalize_document_no
    db = get_db()

    # Step 1: Get all docs (include gpi_folder_correct to know current state)
    all_docs = await db[DOCS_COLL].find({}, {"_id": 1, "run_id": 1, "document_id": 1,
        "file_name": 1, "gpi_doc_type": 1, "gpi_vendor": 1, "gpi_po": 1,
        "gpi_folder_output": 1, "gpi_folder_correct": 1,
        "folder_truth": 1, "is_international": 1}).to_list(5000)

    # Step 2: Collect unique PO numbers and check which exist in BC
    po_set = set()
    for d in all_docs:
        po = d.get("gpi_po", "").strip()
        if po:
            po_set.add(po)

    po_resolved_map = {}  # {po_number: True/False}
    cache_service = get_cache_service()
    if cache_service and po_set:
        from services.bc_reference_cache_service import normalize_document_no
        for po in po_set:
            try:
                # S9 workflow: check if PO exists as a BC purchase order
                # Check bc_document_no, bc_external_document_no, and normalized forms
                norm = normalize_document_no(po)
                match = await cache_service.collection.find_one(
                    {
                        "bc_entity_type": "purchase_order",
                        "$or": [
                            {"bc_document_no": po},
                            {"bc_external_document_no": po},
                            {"normalized_document_no": norm},
                            {"normalized_external_ref": norm},
                        ]
                    },
                    {"_id": 0, "bc_document_no": 1}
                )
                po_resolved_map[po] = match is not None
            except Exception:
                pass  # leave as unknown (won't affect routing)

    resolved_count = sum(1 for v in po_resolved_map.values() if v)
    unresolved_pos = [po for po, v in po_resolved_map.items() if not v]

    # Step 3: Re-route docs, only apply changes that IMPROVE accuracy
    updated_count = 0
    changes = []  # detailed log of what changed
    made_worse = []  # track any regressions
    
    for d in all_docs:
        doc_id = d["_id"]
        po = d.get("gpi_po", "").strip()
        old_folder = d.get("gpi_folder_output", "")
        old_correct = d.get("gpi_folder_correct")
        po_resolved = po_resolved_map.get(po)  # True, False, or None

        sim_doc = {
            "file_name": d.get("file_name", ""),
            "document_type": d.get("gpi_doc_type", ""),
            "vendor_canonical": d.get("gpi_vendor", ""),
            "extracted_fields": {
                "order_number": po,
                "po_number": po,
            },
            "is_international": d.get("is_international", False),
            "bc_po_resolved": po_resolved,
        }

        try:
            is_intl = d.get("is_international", False)
            folder_path, reason, _ = determine_folder_path(sim_doc, is_international=is_intl)
        except Exception:
            folder_path = old_folder
            reason = "error"

        truth = d.get("folder_truth", "")
        new_correct = _folders_match(truth, folder_path) if truth else None

        # Safety: skip if this would break a currently-correct doc
        if old_correct is True and new_correct is False:
            made_worse.append({
                "file_name": d.get("file_name", "")[:60],
                "po": po,
                "old_folder": old_folder,
                "new_folder": folder_path,
                "truth": truth,
                "reason": reason,
                "bc_po_resolved": po_resolved,
            })
            if not dry_run:
                # DON'T apply — keep the old correct folder
                continue

        if folder_path != old_folder:
            changes.append({
                "file_name": d.get("file_name", "")[:60],
                "po": po,
                "old_folder": old_folder,
                "new_folder": folder_path,
                "truth": truth,
                "was_correct": old_correct,
                "now_correct": new_correct,
                "reason": reason,
                "bc_po_resolved": po_resolved,
            })

        if not dry_run:
            update_set = {
                "gpi_folder_output": folder_path,
                "gpi_folder_correct": new_correct,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
            if po_resolved is not None:
                update_set["bc_po_resolved"] = po_resolved
            await db[DOCS_COLL].update_one({"_id": doc_id}, {"$set": update_set})
            if folder_path != old_folder:
                updated_count += 1

    if not dry_run:
        remaining = await db[DOCS_COLL].count_documents({"gpi_folder_correct": False})
    else:
        # Simulate remaining: current mismatches - fixes + regressions
        current_mismatches = sum(1 for d in all_docs if d.get("gpi_folder_correct") is False)
        fixes = sum(1 for c in changes if c.get("was_correct") is False and c.get("now_correct") is True)
        remaining = current_mismatches - fixes

    return {
        "dry_run": dry_run,
        "total_docs": len(all_docs),
        "unique_pos": len(po_set),
        "pos_resolved_in_bc": resolved_count,
        "pos_not_in_bc": len(unresolved_pos),
        "unresolved_pos_sample": unresolved_pos[:20],
        "folders_changed": updated_count if not dry_run else len(changes),
        "would_regress_count": len(made_worse),
        "would_regress_sample": made_worse[:10],
        "remaining_mismatches": remaining,
        "changes_sample": changes[:20],
    }




@router.get("/runs/{run_id}/metrics")
async def get_run_metrics(run_id: str):
    db = get_db()
    docs = await db[DOCS_COLL].find({"run_id": run_id}, {"_id": 0}).to_list(2000)
    metrics = _calc_metrics(docs)
    breakdowns = _calc_breakdowns(docs)
    return {"metrics": metrics, "breakdowns": breakdowns}


@router.get("/runs/{run_id}/folder-alignment")
async def get_folder_alignment(run_id: str):
    """
    Folder Alignment Report: compares S9 output folders vs GPI Hub folder structure.
    
    For each folder appearing in the benchmark data, shows:
    - S9 folder path and doc count
    - Matching GPI Hub folder (if any)
    - GPI routing accuracy for that folder
    - Gaps: S9 folders with no GPI match
    """
    from services.folder_routing_service import FOLDER_STRUCTURE, get_all_folder_paths

    db = get_db()
    docs = await db[DOCS_COLL].find({"run_id": run_id}, {"_id": 0}).to_list(2000)
    if not docs:
        raise HTTPException(404, "No documents in this run")

    gpi_folders = set(get_all_folder_paths())

    # Collect unique S9 folders and GPI folders from the data
    s9_folder_docs = {}  # s9_folder -> list of docs
    gpi_folder_docs = {}  # gpi_folder -> list of docs
    
    for d in docs:
        s9f = d.get("s9_folder_output", "").strip()
        gpif = d.get("gpi_folder_output", "").strip()
        
        if s9f:
            s9_folder_docs.setdefault(s9f, []).append(d)
        if gpif:
            # Normalize: strip PO subfolder for comparison
            base_gpif = gpif.split("/")[0] if "/" in gpif else gpif
            gpi_folder_docs.setdefault(base_gpif, []).append(d)

    def _folder_match(s9_path: str) -> dict:
        """Check if S9 folder has a corresponding GPI Hub folder."""
        s9_lower = s9_path.lower().strip().rstrip("/")
        
        # Exact match
        for gf in gpi_folders:
            if gf.lower() == s9_lower:
                return {"match": "exact", "gpi_folder": gf}
        
        # Partial/contains match
        for gf in gpi_folders:
            gf_lower = gf.lower()
            if s9_lower in gf_lower or gf_lower in s9_lower:
                return {"match": "partial", "gpi_folder": gf}
        
        # Word overlap match
        s9_words = set(re.split(r'[\s/\\-]+', s9_lower))
        best_overlap = 0
        best_folder = None
        for gf in gpi_folders:
            gf_words = set(re.split(r'[\s/\\-]+', gf.lower()))
            overlap = len(s9_words & gf_words)
            if overlap > best_overlap and overlap >= 2:
                best_overlap = overlap
                best_folder = gf
        if best_folder:
            return {"match": "fuzzy", "gpi_folder": best_folder}
        
        return {"match": "none", "gpi_folder": None}

    # Build alignment report
    alignment = []
    for s9_folder, folder_docs in sorted(s9_folder_docs.items(), key=lambda x: -len(x[1])):
        match = _folder_match(s9_folder)
        
        # Count folder accuracy for this S9 folder
        gpi_correct = sum(1 for d in folder_docs if d.get("gpi_folder_correct"))
        s9_correct = sum(1 for d in folder_docs if d.get("s9_folder_correct"))
        
        # What GPI actually routes these docs to
        gpi_destinations = {}
        for d in folder_docs:
            dest = d.get("gpi_folder_output", "")
            if dest:
                base = dest.split("/")[0] if "/" in dest else dest
                gpi_destinations[base] = gpi_destinations.get(base, 0) + 1
        
        alignment.append({
            "s9_folder": s9_folder,
            "doc_count": len(folder_docs),
            "gpi_match": match["match"],
            "gpi_matched_folder": match["gpi_folder"],
            "gpi_folder_correct": gpi_correct,
            "s9_folder_correct": s9_correct,
            "gpi_routing_destinations": gpi_destinations,
        })

    # Identify GPI folders not seen in S9 data
    s9_folder_set = set(s9_folder_docs.keys())
    gpi_only_folders = []
    for gf in sorted(gpi_folders):
        found = False
        for s9f in s9_folder_set:
            if gf.lower() in s9f.lower() or s9f.lower() in gf.lower():
                found = True
                break
        if not found:
            gpi_only_folders.append(gf)

    # Summary stats
    total_s9_folders = len(s9_folder_docs)
    matched = sum(1 for a in alignment if a["gpi_match"] != "none")
    gaps = sum(1 for a in alignment if a["gpi_match"] == "none")

    return {
        "summary": {
            "total_s9_folders": total_s9_folders,
            "matched_to_gpi": matched,
            "gaps": gaps,
            "gpi_total_folders": len(gpi_folders),
            "gpi_only_folders": len(gpi_only_folders),
        },
        "alignment": alignment,
        "gpi_only_folders": gpi_only_folders,
    }



# ═══════════════════════════════════════════════════════════════
# AUTO-POST READINESS ANALYSIS
# ═══════════════════════════════════════════════════════════════

@router.get("/runs/{run_id}/auto-post-readiness")
async def get_auto_post_readiness(run_id: str):
    """
    Analyze what % of AP invoices in a benchmark run could auto-create
    a purchase invoice in BC today.
    
    Checks each AP invoice against the auto-post criteria:
    1. Document classified as AP_Invoice
    2. Vendor matched to BC (vendor_canonical present)
    3. Invoice number extracted
    4. Invoice date extracted
    5. Amount extracted
    6. PO number extracted (for matching to BC purchase order)
    7. Document in SharePoint
    
    Cross-references hub_documents for full extraction data.
    """
    db = get_db()
    docs = await db[DOCS_COLL].find({"run_id": run_id}, {"_id": 0}).to_list(2000)
    if not docs:
        raise HTTPException(404, "No documents in this run")

    # Filter to AP invoices only
    ap_invoices = [d for d in docs if (d.get("gpi_doc_type") or d.get("doc_type_truth") or "").upper().replace(" ", "_") in ("AP_INVOICE",)]
    non_ap = len(docs) - len(ap_invoices)

    results = []
    ready_count = 0
    blockers_summary = {}

    for doc in ap_invoices:
        fname = doc.get("file_name", "")
        vendor = doc.get("gpi_vendor", "")
        po = doc.get("gpi_po", "")
        amount = doc.get("gpi_amount")
        
        # Try to get richer data from hub_documents
        hub_doc = None
        doc_id_from_name = ""
        id_m = re.match(r'^(\d{4,})', fname)
        if id_m:
            doc_id_from_name = id_m.group(1)
        hub_doc = await _find_hub_document(db, doc_id_from_name, fname)

        # Build criteria checklist
        criteria = {}
        blockers = []

        # 1. Vendor matched to BC
        vendor_canonical = None
        vendor_match_method = ""
        if hub_doc:
            vendor_canonical = hub_doc.get("vendor_canonical") or hub_doc.get("vendor_id") or hub_doc.get("vendor_no")
            vendor_match_method = hub_doc.get("vendor_match_method") or hub_doc.get("match_method") or ""
            # Also check bc_vendor_no from validation results
            if not vendor_canonical:
                vr = hub_doc.get("validation_results") or {}
                vendor_canonical = vr.get("bc_vendor_no") or vr.get("vendor_no")
            # Also check vendor match details
            if not vendor_canonical:
                vmd = hub_doc.get("vendor_match_details") or {}
                vendor_canonical = vmd.get("bc_vendor_no") or vmd.get("vendor_no")

        # ACTIVE RESOLUTION: If vendor name exists but no BC match stored,
        # run the vendor alias/BC cache lookup to try resolving it now.
        # This catches docs that were ingested before vendor matching was wired,
        # or where the original match failed but aliases/BC cache have since been updated.
        if not vendor_canonical and vendor:
            try:
                from services.vendor_matching import lookup_vendor_alias
                vendor_normalized = vendor.strip().upper()
                alias_result = await lookup_vendor_alias(vendor_normalized)
                if alias_result and alias_result.get("vendor_canonical"):
                    resolved_method = alias_result.get("vendor_match_method", "benchmark_resolve")
                    # Only accept high-confidence matches (not fuzzy_candidate)
                    if resolved_method not in ("fuzzy_candidate",):
                        vendor_canonical = alias_result["vendor_canonical"]
                        vendor_match_method = f"benchmark:{resolved_method}"
                        logger.info("Benchmark resolved vendor '%s' -> %s via %s",
                                    vendor, vendor_canonical, resolved_method)
            except Exception as e:
                logger.debug("Benchmark vendor resolution failed for '%s': %s", vendor, str(e))

        # Fallback: check vendor_aliases and hub_bc_vendors directly by name
        if not vendor_canonical and vendor:
            try:
                # Check vendor_aliases for this vendor name
                alias_doc = await db.vendor_aliases.find_one({
                    "$or": [
                        {"vendor_name": {"$regex": f"^{re.escape(vendor)}$", "$options": "i"}},
                        {"alias_string": {"$regex": f"^{re.escape(vendor)}$", "$options": "i"}},
                        {"normalized": vendor.strip().upper()},
                    ]
                }, {"_id": 0, "canonical_vendor_id": 1, "vendor_no": 1})
                if alias_doc:
                    vendor_canonical = alias_doc.get("canonical_vendor_id") or alias_doc.get("vendor_no")
                    vendor_match_method = "benchmark:alias_direct"
            except Exception:
                pass

        # Last resort: case-insensitive name match against cached BC vendors
        if not vendor_canonical and vendor:
            try:
                bc_v = await db.hub_bc_vendors.find_one({
                    "$or": [
                        {"displayName": {"$regex": f"^{re.escape(vendor)}$", "$options": "i"}},
                        {"name_normalized": vendor.strip().upper()},
                    ]
                }, {"_id": 0, "number": 1, "displayName": 1})
                if bc_v:
                    vendor_canonical = bc_v.get("number")
                    vendor_match_method = "benchmark:bc_cache_name"
                    logger.info("Benchmark resolved vendor '%s' -> %s via bc_cache_name", vendor, vendor_canonical)
            except Exception:
                pass

        criteria["vendor_matched_bc"] = bool(vendor_canonical)
        if not vendor_canonical:
            if vendor:
                blockers.append("vendor_name_only")  # Have name but not BC match
            else:
                blockers.append("no_vendor")

        # 2. Invoice number
        invoice_no = None
        if hub_doc:
            invoice_no = (
                hub_doc.get("invoice_number_clean")
                or (hub_doc.get("extracted_fields") or {}).get("invoice_number")
                or (hub_doc.get("ai_extraction") or {}).get("invoice_number")
            )
        criteria["invoice_number"] = bool(invoice_no)
        if not invoice_no:
            blockers.append("no_invoice_number")

        # 3. Invoice date
        invoice_date = None
        if hub_doc:
            invoice_date = (
                hub_doc.get("invoice_date")
                or (hub_doc.get("extracted_fields") or {}).get("invoice_date")
                or (hub_doc.get("ai_extraction") or {}).get("invoice_date")
            )
        criteria["invoice_date"] = bool(invoice_date)
        if not invoice_date:
            blockers.append("no_invoice_date")

        # 4. Amount
        has_amount = amount is not None
        if not has_amount and hub_doc:
            for k in ["amount_float", "total_amount", "invoice_amount"]:
                if hub_doc.get(k):
                    has_amount = True
                    break
            if not has_amount:
                ef = hub_doc.get("extracted_fields") or {}
                ai = hub_doc.get("ai_extraction") or {}
                has_amount = bool(ef.get("amount") or ef.get("total_amount") or ai.get("total_amount"))
        criteria["amount"] = has_amount
        if not has_amount:
            blockers.append("no_amount")

        # 5. PO number (for BC order matching)
        has_po = bool(po)
        if not has_po and hub_doc:
            has_po = bool(
                hub_doc.get("po_number_extracted") or hub_doc.get("po_number")
                or (hub_doc.get("extracted_fields") or {}).get("po_number")
                or (hub_doc.get("ai_extraction") or {}).get("po_number")
                or (hub_doc.get("normalized_fields") or {}).get("po_number")
                or hub_doc.get("bc_po_number")
                or hub_doc.get("purchase_order_number")
            )
            # Also check if there's a linked BC order already
            if not has_po:
                has_po = bool(
                    hub_doc.get("bc_purchase_order_id") or hub_doc.get("bc_order_number")
                    or hub_doc.get("linked_bc_record_id")
                )
        criteria["po_number"] = has_po
        if not has_po:
            blockers.append("no_po")

        # 6. In SharePoint
        in_sp = False
        if hub_doc:
            in_sp = bool(hub_doc.get("sharepoint_share_link_url") or hub_doc.get("sharepoint_web_url"))
        criteria["in_sharepoint"] = in_sp
        if not in_sp:
            blockers.append("not_in_sharepoint")

        # 7. Confidence — BLENDED with stable vendor score (the feedback loop)
        raw_confidence = 0
        stable_score = 0
        stable_flag = False
        if hub_doc:
            raw_confidence = (
                (hub_doc.get("ai_extraction") or {}).get("confidence", 0) 
                or hub_doc.get("classification_confidence", 0)
                or hub_doc.get("ai_confidence") or 0
            )
            stable_score = hub_doc.get("stable_vendor_score", 0) or 0
            stable_flag = hub_doc.get("stable_vendor_flag", False)

        # If stable data not on hub_doc, look up vendor_intelligence_profiles
        if not stable_flag and vendor_canonical:
            vip = await db.vendor_intelligence_profiles.find_one(
                {"$or": [
                    {"vendor_no": vendor_canonical},
                    {"vendor_no": {"$regex": f"^{re.escape(vendor_canonical)}$", "$options": "i"}},
                ]},
                {"_id": 0, "stable_vendor_flag": 1, "stable_vendor_score": 1}
            )
            if vip:
                stable_flag = vip.get("stable_vendor_flag", False)
                stable_score = vip.get("stable_vendor_score", 0) or 0

        # Also try by vendor name if vendor_canonical didn't match
        if not stable_flag and vendor:
            vip2 = await db.vendor_intelligence_profiles.find_one(
                {"$or": [
                    {"vendor_name": {"$regex": f"^{re.escape(vendor)}$", "$options": "i"}},
                    {"display_name": {"$regex": f"^{re.escape(vendor)}$", "$options": "i"}},
                ]},
                {"_id": 0, "stable_vendor_flag": 1, "stable_vendor_score": 1}
            )
            if vip2:
                stable_flag = vip2.get("stable_vendor_flag", False)
                stable_score = vip2.get("stable_vendor_score", 0) or 0

        # THE FEEDBACK LOOP: stable vendor flag = earned trust
        # Rule: raw confidence alone meeting threshold always passes
        # Stable score can only BOOST, never reduce effective confidence
        if stable_flag and stable_score >= 0.85:
            effective_confidence = max(raw_confidence, stable_score)
        elif stable_score > 0:
            blended = (raw_confidence * 0.4) + (stable_score * 0.6)
            effective_confidence = max(raw_confidence, blended)
        else:
            effective_confidence = raw_confidence

        threshold = 0.90
        criteria["confidence_ok"] = effective_confidence >= threshold
        # Informational fields — NOT used for readiness check
        info = {
            "stable_vendor": stable_flag,
            "stable_score": round(stable_score, 3),
        }
        if effective_confidence < threshold and hub_doc:
            if stable_flag:
                blockers.append(f"low_confidence (stable vendor but score={stable_score:.2f})")
            elif stable_score > 0:
                blockers.append(f"low_confidence (building history: raw={raw_confidence:.2f} stable={stable_score:.2f})")
            else:
                blockers.append("low_confidence (no vendor history)")

        # Overall readiness — only check actual pass/fail criteria
        is_ready = all(criteria.values())
        if is_ready:
            ready_count += 1

        # Count blockers
        for b in blockers:
            blockers_summary[b] = blockers_summary.get(b, 0) + 1

        results.append({
            "file_name": fname,
            "vendor": vendor,
            "vendor_canonical": vendor_canonical,
            "vendor_match_method": vendor_match_method,
            "criteria": {**criteria, **info},
            "blockers": blockers,
            "ready": is_ready,
            "hub_doc_found": hub_doc is not None,
        })

    # Calculate readiness tiers
    tier_one_blocker = sum(1 for r in results if len(r["blockers"]) == 1)
    tier_multiple = sum(1 for r in results if len(r["blockers"]) > 1)
    tier_no_hub = sum(1 for r in results if not r["hub_doc_found"])

    total_ap = len(ap_invoices)
    pct_ready = round(ready_count / total_ap * 100, 1) if total_ap else 0

    # Count resolution methods for visibility
    resolution_methods = {}
    for r in results:
        method = r.get("vendor_match_method", "none")
        if method:
            resolution_methods[method] = resolution_methods.get(method, 0) + 1

    return {
        "summary": {
            "total_documents": len(docs),
            "ap_invoices": total_ap,
            "non_ap_documents": non_ap,
            "auto_post_ready": ready_count,
            "auto_post_ready_pct": pct_ready,
            "one_blocker_away": tier_one_blocker,
            "multiple_blockers": tier_multiple,
            "no_hub_data": tier_no_hub,
        },
        "blockers_distribution": dict(sorted(blockers_summary.items(), key=lambda x: -x[1])),
        "vendor_resolution_methods": dict(sorted(resolution_methods.items(), key=lambda x: -x[1])),
        "criteria_pass_rates": {
            "vendor_matched_bc": round(sum(1 for r in results if r["criteria"]["vendor_matched_bc"]) / total_ap * 100, 1) if total_ap else 0,
            "invoice_number": round(sum(1 for r in results if r["criteria"]["invoice_number"]) / total_ap * 100, 1) if total_ap else 0,
            "invoice_date": round(sum(1 for r in results if r["criteria"]["invoice_date"]) / total_ap * 100, 1) if total_ap else 0,
            "amount": round(sum(1 for r in results if r["criteria"]["amount"]) / total_ap * 100, 1) if total_ap else 0,
            "po_number": round(sum(1 for r in results if r["criteria"]["po_number"]) / total_ap * 100, 1) if total_ap else 0,
            "in_sharepoint": round(sum(1 for r in results if r["criteria"]["in_sharepoint"]) / total_ap * 100, 1) if total_ap else 0,
        },
        "documents": results,
    }

# ═══════════════════════════════════════════════════════════════
# EXPORT (Excel)
# ═══════════════════════════════════════════════════════════════

@router.get("/runs/{run_id}/export")
async def export_run(run_id: str):
    """Export run as Excel with two sheets: Documents + Summary."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    db = get_db()
    run = await db[RUNS_COLL].find_one({"run_id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(404, "Run not found")

    docs = await db[DOCS_COLL].find({"run_id": run_id}, {"_id": 0}).to_list(2000)
    metrics = _calc_metrics(docs)

    wb = Workbook()

    # --- Documents Sheet ---
    ws = wb.active
    ws.title = "Documents"
    headers = [
        "Document ID", "File Name", "Doc Type Truth", "Vendor Truth", "Amount Truth", "PO Truth", "Folder Truth",
        "GPI Ingested", "GPI Doc Type", "GPI Vendor", "GPI Amount", "GPI PO", "GPI Folder",
        "GPI DocType OK", "GPI Vendor OK", "GPI Amount OK", "GPI PO OK", "GPI Folder OK",
        "GPI Needs Review", "GPI Final Status", "GPI Why Wrong",
        "S9 Ingested", "S9 Doc Type", "S9 Vendor", "S9 Amount", "S9 PO", "S9 Folder",
        "S9 DocType OK", "S9 Vendor OK", "S9 Amount OK", "S9 PO OK", "S9 Folder OK",
        "S9 Needs Review", "S9 Final Status", "S9 Why Wrong",
    ]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, d in enumerate(docs, 2):
        vals = [
            d.get("document_id"), d.get("file_name"),
            d.get("doc_type_truth"), d.get("vendor_truth"), d.get("amount_truth"), d.get("po_truth"), d.get("folder_truth"),
            d.get("gpi_ingested"), d.get("gpi_doc_type"), d.get("gpi_vendor"), d.get("gpi_amount"), d.get("gpi_po"), d.get("gpi_folder_output"),
            d.get("gpi_doc_type_correct"), d.get("gpi_vendor_correct"), d.get("gpi_amount_correct"), d.get("gpi_po_correct"), d.get("gpi_folder_correct"),
            d.get("gpi_needs_review"), d.get("gpi_final_status"), ", ".join(d.get("gpi_why_wrong_tags") or []),
            d.get("s9_ingested"), d.get("s9_doc_type"), d.get("s9_vendor"), d.get("s9_amount"), d.get("s9_po"), d.get("s9_folder_output"),
            d.get("s9_doc_type_correct"), d.get("s9_vendor_correct"), d.get("s9_amount_correct"), d.get("s9_po_correct"), d.get("s9_folder_correct"),
            d.get("s9_needs_review"), d.get("s9_final_status"), ", ".join(d.get("s9_why_wrong_tags") or []),
        ]
        for col, v in enumerate(vals, 1):
            ws.cell(row=i, column=col, value=v)

    # --- Summary Sheet ---
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value=f"Bake-Off: {run.get('name', '')}").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"Date: {run.get('test_date', '')}")
    ws2.cell(row=3, column=1, value=f"Documents: {metrics.get('total', 0)}")

    kpi_headers = ["KPI", "GPI Hub", "Square 9", "Delta"]
    for col, h in enumerate(kpi_headers, 1):
        cell = ws2.cell(row=5, column=col, value=h)
        cell.font = Font(bold=True)

    gpi = metrics.get("gpi", {})
    s9 = metrics.get("s9", {})
    kpis = [
        ("Ingest Rate %", gpi.get("ingest_rate"), s9.get("ingest_rate")),
        ("Classification Accuracy %", gpi.get("classification_accuracy"), s9.get("classification_accuracy")),
        ("Vendor Accuracy %", gpi.get("vendor_accuracy"), s9.get("vendor_accuracy")),
        ("Amount Accuracy %", gpi.get("amount_accuracy"), s9.get("amount_accuracy")),
        ("PO Accuracy %", gpi.get("po_accuracy"), s9.get("po_accuracy")),
        ("Folder Accuracy %", gpi.get("folder_accuracy"), s9.get("folder_accuracy")),
        ("No-Touch Rate %", gpi.get("no_touch_rate"), s9.get("no_touch_rate")),
        ("Usable Output Rate %", gpi.get("usable_output_rate"), s9.get("usable_output_rate")),
    ]
    for i, (label, g, s) in enumerate(kpis, 6):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=g)
        ws2.cell(row=i, column=3, value=s)
        if g is not None and s is not None:
            ws2.cell(row=i, column=4, value=round(g - s, 1))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"bakeoff_{run_id}_{run.get('name', 'export').replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/why-wrong-tags")
async def get_why_wrong_tags():
    return {"tags": WHY_WRONG_TAGS}



# ═══════════════════════════════════════════════════════════════
# SHAREPOINT FOLDER SCAN — S9 vs GPI "Folder Diff" Builder
# ═══════════════════════════════════════════════════════════════

class ScanSharePointReq(BaseModel):
    """Configuration for scanning Square 9 output folders on SharePoint."""
    s9_root_folder: str = ""  # Root folder where S9 files documents. Empty = use library root.
    folder_names: Optional[List[str]] = None  # Specific folders to scan. None = scan all known folders.
    max_files_per_folder: int = 50  # Limit per folder to avoid huge scans.
    include_subfolders: bool = True  # Recurse into subfolders.
    auto_match_gpi: bool = True  # Try to match files against hub_documents.
    auto_route_gpi: bool = True  # Run GPI routing logic on matched documents.


async def _resolve_sp_drive():
    """Resolve SharePoint site_id and drive_id. Returns (drive_id, error_msg)."""
    import os, httpx
    from services.sharepoint_service import (
        DEMO_MODE, GRAPH_CLIENT_ID, SHAREPOINT_SITE_HOSTNAME,
        SHAREPOINT_SITE_PATH, SHAREPOINT_LIBRARY_NAME,
    )

    if DEMO_MODE or not GRAPH_CLIENT_ID:
        return None, "DEMO_MODE"

    try:
        from services.sharepoint_service import _get_graph_token
        token = await _get_graph_token()
    except Exception as e:
        logger.warning("Graph token failed, falling back to DEMO_MODE: %s", str(e)[:200])
        return None, "DEMO_MODE"

    async with httpx.AsyncClient(timeout=30.0) as c:
        site_resp = await c.get(
            f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_HOSTNAME}:{SHAREPOINT_SITE_PATH}:",
            headers={"Authorization": f"Bearer {token}"},
        )
        if site_resp.status_code != 200:
            return None, f"Site resolve failed (HTTP {site_resp.status_code})"
        site_id = site_resp.json()["id"]

        drives_resp = await c.get(
            f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives",
            headers={"Authorization": f"Bearer {token}"},
        )
        drives = drives_resp.json().get("value", [])
        lib = SHAREPOINT_LIBRARY_NAME
        drive = next((d for d in drives if d["name"].lower() == lib.lower()), None)
        if not drive:
            drive = next((d for d in drives if d.get("driveType") == "documentLibrary"), None)
        if not drive:
            return None, f"Drive '{lib}' not found. Available: {[d['name'] for d in drives]}"
        return drive["id"], None


async def _list_folder_files(drive_id: str, folder_path: str, max_files: int = 50, recurse: bool = True):
    """List files in a SharePoint folder via Graph API. Returns list of (file_name, folder_path, file_metadata)."""
    import httpx
    from services.sharepoint_service import _get_graph_token

    token = await _get_graph_token()
    files = []

    async with httpx.AsyncClient(timeout=30.0) as c:
        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{folder_path}:/children?$top={max_files}"
        resp = await c.get(url, headers={"Authorization": f"Bearer {token}"})

        if resp.status_code == 404:
            return []  # Folder doesn't exist
        if resp.status_code != 200:
            logger.warning("Failed to list %s: HTTP %s", folder_path, resp.status_code)
            return []

        items = resp.json().get("value", [])
        for item in items:
            if "file" in item:
                files.append({
                    "file_name": item["name"],
                    "s9_folder": folder_path,
                    "size": item.get("size", 0),
                    "created": item.get("createdDateTime", ""),
                    "modified": item.get("lastModifiedDateTime", ""),
                    "web_url": item.get("webUrl", ""),
                })
            elif "folder" in item and recurse:
                sub_path = f"{folder_path}/{item['name']}"
                sub_files = await _list_folder_files(drive_id, sub_path, max_files, recurse=True)
                files.extend(sub_files)

            if len(files) >= max_files:
                break

    return files[:max_files]


def _generate_demo_files(folder_names, max_per_folder):
    """Generate realistic demo file listings for DEMO_MODE preview testing."""
    from services.folder_routing_service import FOLDER_STRUCTURE
    demo_vendors = {
        "Dropship International Documents": [
            ("112148_Fevisa_ML179859_031192026.pdf", "FEVISA INDUSTRIAL S.A. DE C.V."),
            ("112201_Vitrocrisa_ML180021_031192026.pdf", "VITROCRISA S.A. DE C.V."),
            ("112305_Envases_PO44821_031192026.pdf", "ENVASES UNIVERSALES DE MEXICO"),
        ],
        "Dropship Not International Documents": [
            ("112089_Ball_PO88432_031192026.pdf", "BALL CORPORATION"),
            ("112095_Anchor_PO88501_031192026.pdf", "ANCHOR GLASS CONTAINER"),
            ("112102_OI_PO88612_031192026.pdf", "OWENS-ILLINOIS INC"),
        ],
        "Freight Issues": [
            ("FRT_XPO_BOL8834521_031192026.pdf", "XPO LOGISTICS"),
            ("FRT_SAIA_BOL7721034_031192026.pdf", "SAIA INC"),
        ],
        "Warehouse International Documents": [
            ("WH_112411_Fevisa_ML180102_031192026.pdf", "FEVISA INDUSTRIAL S.A. DE C.V."),
        ],
        "Warehouse Not International Documents": [
            ("WH_112320_Ball_PO88701_031192026.pdf", "BALL CORPORATION"),
            ("WH_112321_GTs_PO88702_031192026.pdf", "GT'S LIVING FOODS"),
        ],
        "Vendor Credit Memos": [
            ("CM_Ball_Dunnage_112501_031192026.pdf", "BALL CORPORATION"),
        ],
        "Tooling Invoices": [
            ("TOOL_112601_Ball_Mold_031192026.pdf", "BALL CORPORATION"),
        ],
        "S&H Invoices Approved Documents": [
            ("SH_112701_Warehouse_031192026.pdf", "GAMER PACKAGING WAREHOUSE"),
        ],
    }

    folders_to_scan = folder_names or list(FOLDER_STRUCTURE.values())
    if not folder_names:
        folders_to_scan = [v["path"] for v in FOLDER_STRUCTURE.values()]

    files = []
    for folder_path in folders_to_scan:
        vendor_files = demo_vendors.get(folder_path, [])
        for fname, vendor in vendor_files[:max_per_folder]:
            files.append({
                "file_name": fname,
                "s9_folder": folder_path,
                "size": 245000,
                "created": "2026-03-19T10:30:00Z",
                "modified": "2026-03-19T10:30:00Z",
                "web_url": f"https://gamerpackaging.sharepoint.com/sites/GPI-DocumentHub-Test/Shared Documents/{folder_path}/{fname}",
                "demo_vendor": vendor,
            })
    return files


@router.post("/runs/{run_id}/scan-sharepoint")
async def scan_sharepoint_folders(run_id: str, req: ScanSharePointReq):
    """
    Scan Square 9 output folders on SharePoint and auto-create benchmark entries.

    For each file found:
    1. The S9 folder path IS the S9 classification
    2. Match against hub_documents for GPI Hub extraction data
    3. Run GPI routing logic to determine where GPI Hub would file it
    4. Create a benchmark document entry with both sides pre-filled
    """
    db = get_db()
    run = await db[RUNS_COLL].find_one({"run_id": run_id})
    if not run:
        raise HTTPException(404, "Run not found")

    from services.folder_routing_service import FOLDER_STRUCTURE, determine_folder_path

    # Determine which folders to scan
    if req.folder_names:
        folders_to_scan = req.folder_names
    else:
        folders_to_scan = [v["path"] for v in FOLDER_STRUCTURE.values()]

    # Prepend S9 root if provided
    if req.s9_root_folder:
        folders_to_scan = [f"{req.s9_root_folder}/{f}" for f in folders_to_scan]

    # Resolve SharePoint drive
    drive_id, err = await _resolve_sp_drive()

    all_files = []
    is_demo = err == "DEMO_MODE"

    if is_demo:
        # In DEMO_MODE, generate realistic test data
        raw_folders = req.folder_names or [v["path"] for v in FOLDER_STRUCTURE.values()]
        all_files = _generate_demo_files(raw_folders, req.max_files_per_folder)
        logger.info("[Benchmark SP Scan] DEMO_MODE: generated %d mock files", len(all_files))
    elif err:
        raise HTTPException(502, f"SharePoint connection failed: {err}")
    else:
        # Live SharePoint scan
        for folder_path in folders_to_scan:
            files = await _list_folder_files(
                drive_id, folder_path,
                max_files=req.max_files_per_folder,
                recurse=req.include_subfolders,
            )
            all_files.extend(files)
            logger.info("[Benchmark SP Scan] %s: %d files", folder_path, len(files))

    if not all_files:
        return {"scanned_folders": len(folders_to_scan), "files_found": 0, "documents_created": 0,
                "message": "No files found in scanned folders"}

    # Deduplicate by filename (same file may appear in subfolders)
    seen = set()
    unique_files = []
    for f in all_files:
        if f["file_name"] not in seen:
            seen.add(f["file_name"])
            unique_files.append(f)

    # Create benchmark document entries
    created = 0
    matched_gpi = 0
    routed_gpi = 0

    for file_info in unique_files:
        fname = file_info["file_name"]
        s9_folder = file_info["s9_folder"]

        # Check if doc already exists in this run
        existing = await db[DOCS_COLL].find_one({"run_id": run_id, "file_name": fname})
        if existing:
            continue

        # Build base document
        doc_data = {
            "file_name": fname,
            "source_path": file_info.get("web_url", s9_folder),
            "received_date": (file_info.get("created") or "")[:10],
            "folder_truth": s9_folder,  # S9's folder IS the ground truth for folder routing
        }

        # Use demo_vendor hint if available
        demo_vendor = file_info.get("demo_vendor", "")

        doc = _make_doc(run_id, doc_data)

        # S9 fields — the folder IS the classification
        doc["s9_ingested"] = True
        doc["s9_folder_output"] = s9_folder

        # Try to match against hub_documents for GPI data (improved multi-strategy)
        hub_doc = None
        if req.auto_match_gpi:
            # Extract document_id from filename pattern (e.g., "112148" from "112148_Vendor_...")
            doc_id_from_name = ""
            id_m = re.match(r'^(\d{4,})', fname)
            if id_m:
                doc_id_from_name = id_m.group(1)
            hub_doc = await _find_hub_document(db, doc_id_from_name, fname)

        if hub_doc:
            ef = hub_doc.get("extracted_fields") or {}
            nf = hub_doc.get("normalized_fields") or {}
            ai = hub_doc.get("ai_extraction") or {}
            doc["gpi_ingested"] = True
            doc["gpi_auto_linked"] = True
            doc["gpi_source_document_id"] = hub_doc.get("id") or hub_doc.get("document_id", "")
            doc["gpi_doc_type"] = (
                hub_doc.get("document_type") or hub_doc.get("doc_type")
                or hub_doc.get("suggested_job_type") or ai.get("document_type") or ""
            )
            doc["gpi_vendor"] = (
                hub_doc.get("vendor_canonical") or hub_doc.get("vendor_name")
                or hub_doc.get("vendor_no") or nf.get("vendor") or ef.get("vendor")
                or ai.get("vendor") or ""
            )
            doc["gpi_po"] = (
                hub_doc.get("po_number_extracted") or hub_doc.get("po_number")
                or hub_doc.get("purchase_order_number") or hub_doc.get("bol_number_extracted")
                or nf.get("po_number") or ef.get("po_number") or ai.get("po_number") or ""
            )

            gpi_amount = None
            for amt_src in [
                nf.get("amount"), nf.get("total_amount"), ef.get("amount"), ef.get("total_amount"),
                ai.get("amount"), hub_doc.get("total_amount"), hub_doc.get("invoice_amount"),
            ]:
                if amt_src is not None:
                    try:
                        gpi_amount = float(str(amt_src).replace(",", "").replace("$", ""))
                        break
                    except (ValueError, TypeError):
                        continue
            doc["gpi_amount"] = gpi_amount

            # Seed truth fields from GPI extraction
            if doc["gpi_vendor"]:
                doc["vendor_truth"] = doc["gpi_vendor"]
            if doc["gpi_doc_type"]:
                doc["doc_type_truth"] = doc["gpi_doc_type"]
            if gpi_amount is not None:
                doc["amount_truth"] = gpi_amount
            if doc["gpi_po"]:
                doc["po_truth"] = doc["gpi_po"]

            # Map status
            status = hub_doc.get("status", "")
            if status in ("Completed", "Filed", "Posted", "completed", "filed", "posted"):
                doc["gpi_final_status"] = "Usable"
                doc["gpi_needs_review"] = "None"
            elif status in ("Pending", "In_Review", "captured", "pending", "in_review"):
                doc["gpi_final_status"] = "Partial"
                doc["gpi_needs_review"] = "Minor"

            # Run GPI routing logic
            if req.auto_route_gpi:
                is_intl = hub_doc.get("is_international", False)
                try:
                    folder_path, reason, details = determine_folder_path(hub_doc, is_international=is_intl)
                    doc["gpi_folder_output"] = folder_path
                    doc["gpi_notes"] = f"Auto-routed: {reason}"
                    routed_gpi += 1
                except Exception as e:
                    doc["gpi_notes"] = f"Routing error: {str(e)}"

            matched_gpi += 1
        elif demo_vendor and is_demo:
            # In DEMO_MODE, simulate GPI extraction from filename patterns
            doc["gpi_ingested"] = True
            doc["gpi_vendor"] = demo_vendor

            # Parse PO from filename pattern (e.g., _PO88432_ or _ML179859_)
            po_match = re.search(r'[_]((?:PO|ML|SO)\d+)[_]', fname, re.IGNORECASE)
            if po_match:
                doc["gpi_po"] = po_match.group(1)

            # Infer doc type from filename
            if fname.startswith("CM_"):
                doc["gpi_doc_type"] = "Credit_Memo"
            elif fname.startswith("FRT_"):
                doc["gpi_doc_type"] = "Freight_Document"
            elif fname.startswith("SH_"):
                doc["gpi_doc_type"] = "S&H_Invoice"
            elif fname.startswith("TOOL_"):
                doc["gpi_doc_type"] = "Tooling_Invoice"
            elif fname.startswith("WH_"):
                doc["gpi_doc_type"] = "AP_Invoice"
            else:
                doc["gpi_doc_type"] = "AP_Invoice"

            doc["vendor_truth"] = demo_vendor
            doc["doc_type_truth"] = doc["gpi_doc_type"]

            # Run routing logic with simulated data
            if req.auto_route_gpi:
                sim_doc = {
                    "file_name": fname,
                    "document_type": doc["gpi_doc_type"],
                    "vendor_canonical": demo_vendor,
                    "extracted_fields": {"vendor": demo_vendor, "po_number": doc.get("gpi_po", "")},
                    "normalized_fields": {"vendor": demo_vendor, "po_number": doc.get("gpi_po", "")},
                    "po_number_extracted": doc.get("gpi_po", ""),
                }
                is_intl = "S.A. DE C.V." in demo_vendor.upper() or (
                    "international" in s9_folder.lower() and "not international" not in s9_folder.lower()
                )
                try:
                    folder_path, reason, details = determine_folder_path(sim_doc, is_international=is_intl)
                    doc["gpi_folder_output"] = folder_path
                    doc["gpi_notes"] = f"Demo auto-routed: {reason}"
                    routed_gpi += 1
                except Exception as e:
                    doc["gpi_notes"] = f"Routing error: {str(e)}"

            matched_gpi += 1

        await db[DOCS_COLL].insert_one(doc)
        doc.pop("_id", None)
        created += 1

    # Auto-score all new docs
    all_docs = await db[DOCS_COLL].find({"run_id": run_id}).to_list(2000)
    for d in all_docs:
        d.pop("_id", None)
        scores = auto_score_correctness(d)
        if scores:
            await db[DOCS_COLL].update_one(
                {"run_id": run_id, "doc_uid": d["doc_uid"]},
                {"$set": scores}
            )

    # Update run status
    now = datetime.now(timezone.utc).isoformat()
    await db[RUNS_COLL].update_one(
        {"run_id": run_id},
        {"$set": {"status": "in_progress", "updated_at": now}}
    )

    return {
        "scanned_folders": len(folders_to_scan),
        "files_found": len(unique_files),
        "documents_created": created,
        "gpi_matched": matched_gpi,
        "gpi_routed": routed_gpi,
        "demo_mode": is_demo,
        "message": f"Scanned {len(folders_to_scan)} folders, found {len(unique_files)} files, created {created} benchmark entries"
            + (f" (DEMO_MODE: using simulated data)" if is_demo else ""),
    }
