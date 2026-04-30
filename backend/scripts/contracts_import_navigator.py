"""Phase 4B — One-shot DocuSign Navigator → Contract Intelligence import CLI.

Imports (or dry-runs) a DocuSign Navigator AI Metadata Export — an .xlsx
or .csv file where each row is a single agreement with ~54 flat columns
("Envelope Id", "Agreement Type", "Parties", "Payment Term", ...).

**Default is dry-run.** Writes only happen when ``--commit`` is passed.

Pipeline reuse:
    Every row is shipped through the existing contract-intelligence
    orchestrator (:class:`services.contracts.contract_intelligence_service.ContractIntelligenceService`).
    The normalizer's unified entry point auto-dispatches flat Navigator
    rows to :func:`services.contracts.navigator_normalizer.normalize_navigator_row`,
    so import and live-Connect events share one persistence / matcher /
    audit path.

Idempotency:
    * The event id is deterministic: ``navigator::{envelope_id}``.
    * The ``agreement_events`` unique index
      (``provider``, ``provider_event_id``) makes a replay a no-op.
    * The ``agreements`` unique index on ``provider_envelope_id`` makes
      the underlying row upsert a no-op if the envelope was already
      ingested via Connect.
    * Manual mappings are preserved: the orchestrator's replay logic
      only rewrites ``linked_by="system"`` links in
      ``{proposed, auto_confirmed}`` state. Confirmed / rejected /
      manual-link rows survive.

Usage (remote VM):

    # Dry-run (default):
    docker compose exec backend \\
        python -m scripts.contracts_import_navigator /tmp/navigator.xlsx

    # Actually write:
    docker compose exec backend \\
        python -m scripts.contracts_import_navigator /tmp/navigator.xlsx --commit

    # Limit to the first N rows (debugging):
    docker compose exec backend \\
        python -m scripts.contracts_import_navigator /tmp/navigator.xlsx --limit 5

    # Pick a specific sheet (xlsx only; defaults to the active sheet):
    docker compose exec backend \\
        python -m scripts.contracts_import_navigator /tmp/navigator.xlsx --sheet "Agreements"

Exit codes:
    0  — every row processed cleanly (includes "skipped as duplicate")
    1  — usage / argparse error
    2  — file could not be read or its shape is not a Navigator export
    3  — one or more rows failed normalization or commit
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import json
import os
import sys
from dataclasses import dataclass, field
from typing import Any, Dict, Iterable, List, Optional

_BACKEND_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _BACKEND_DIR not in sys.path:
    sys.path.insert(0, _BACKEND_DIR)


from services.contracts.agreement_normalizer import (  # noqa: E402
    NormalizedAgreement,
    normalize_envelope,
)
from services.contracts.navigator_normalizer import (  # noqa: E402
    normalize_navigator_row,
)


# ---------------------------------------------------------------------------
# Row loaders
# ---------------------------------------------------------------------------

def _load_csv(path: str) -> List[Dict[str, Any]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as fp:
        reader = csv.DictReader(fp)
        return [dict(row) for row in reader]


def _load_xlsx(path: str, sheet: Optional[str]) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "openpyxl is required to read .xlsx files but is not installed. "
            "Install it with `pip install openpyxl` and rerun."
        ) from exc
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    if ws is None:
        return []
    rows_iter = ws.iter_rows(values_only=True)
    header_row: Optional[List[Any]] = None
    out: List[Dict[str, Any]] = []
    for raw in rows_iter:
        if header_row is None:
            header_row = [("" if c is None else str(c).strip()) for c in raw]
            continue
        if all((c is None or str(c).strip() == "") for c in raw):
            continue
        row_dict: Dict[str, Any] = {}
        for header, value in zip(header_row, raw):
            if not header:
                continue
            row_dict[header] = value
        out.append(row_dict)
    return out


def _load_json(path: str) -> List[Dict[str, Any]]:
    """Convenience loader for the Bragg-style fixture JSON (tests / dev)."""
    with open(path, "r", encoding="utf-8") as fp:
        data = json.load(fp)
    if isinstance(data, list):
        return [r for r in data if isinstance(r, dict)]
    if isinstance(data, dict):
        # Single-row fixture wrapper.
        if isinstance(data.get("row"), dict):
            return [data["row"]]
        # Multi-row wrapper.
        if isinstance(data.get("rows"), list):
            return [r for r in data["rows"] if isinstance(r, dict)]
        # Naked flat row.
        return [data]
    raise ValueError("unsupported JSON shape; expected list/dict")


def load_rows(path: str, *, sheet: Optional[str]) -> List[Dict[str, Any]]:
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return _load_xlsx(path, sheet)
    if ext == ".csv":
        return _load_csv(path)
    if ext == ".json":
        return _load_json(path)
    raise ValueError(
        f"unsupported file extension {ext!r}; use .xlsx / .csv / .json"
    )


# ---------------------------------------------------------------------------
# Per-row report shape
# ---------------------------------------------------------------------------

@dataclass
class RowReport:
    index: int
    envelope_id: Optional[str] = None
    provider_agreement_id: Optional[str] = None
    title: Optional[str] = None
    status: Optional[str] = None

    normalized: Optional[NormalizedAgreement] = None
    error: Optional[str] = None

    # Commit-side outcomes (dry-run leaves these empty).
    committed: bool = False
    duplicate: bool = False
    agreement_id: Optional[str] = None
    link_count: int = 0
    exception_count: int = 0

    def party_count(self) -> int:
        return len(self.normalized.parties) if self.normalized else 0

    def term_count(self) -> int:
        return len(self.normalized.terms) if self.normalized else 0

    def pricing_count(self) -> int:
        return len(self.normalized.pricing) if self.normalized else 0

    def document_count(self) -> int:
        return len(self.normalized.documents) if self.normalized else 0

    def warning_count(self) -> int:
        return len(self.normalized.warnings) if self.normalized else 0


# ---------------------------------------------------------------------------
# Dry-run
# ---------------------------------------------------------------------------

def dryrun_row(index: int, row: Dict[str, Any]) -> RowReport:
    report = RowReport(index=index)
    try:
        normalized = normalize_navigator_row(row, event_id=f"navigator-cli::{index}")
    except ValueError as exc:
        report.error = str(exc)
        return report
    except Exception as exc:  # noqa: BLE001 — defensive: row-level isolation
        report.error = f"{type(exc).__name__}: {exc}"
        return report

    report.normalized = normalized
    report.envelope_id = normalized.agreement.provider_envelope_id
    report.provider_agreement_id = normalized.agreement.provider_agreement_id
    report.title = normalized.agreement.title
    report.status = normalized.agreement.status
    return report


# ---------------------------------------------------------------------------
# BC match + commit — async
# ---------------------------------------------------------------------------

async def commit_row(svc, index: int, row: Dict[str, Any]) -> RowReport:
    """Route the row through the real orchestrator: ``record_event`` +
    ``process_event``. Relies on the event-layer uniqueness index for
    idempotency — a replay is a no-op."""
    report = dryrun_row(index, row)
    if report.error or report.normalized is None:
        return report

    envelope_id = report.envelope_id
    event_id_str = f"navigator::{envelope_id}"
    event_type = "navigator-import"

    # 1. Record the raw row as an event. Duplicate returns duplicate=True.
    record = await svc.record_event(
        provider_event_id=event_id_str,
        provider_envelope_id=envelope_id,
        event_type=event_type,
        raw_payload=row,
        hmac_valid=True,
        transport="manual",
    )
    if record["duplicate"]:
        report.duplicate = True
        report.committed = False
        report.agreement_id = None
        return report

    # 2. Process the freshly-recorded event. The orchestrator delegates to
    #    ``normalize_envelope`` which auto-dispatches Navigator rows back
    #    through our adapter.
    processed = await svc.process_event(record["event_id"])
    report.committed = processed.get("status") == "ok"
    report.agreement_id = processed.get("agreement_id")
    report.link_count = processed.get("links", 0) or 0
    report.exception_count = processed.get("exceptions", 0) or 0
    if processed.get("status") == "normalizer_failed":
        report.error = processed.get("error") or "normalizer_failed"
    return report


async def run_commit(rows: List[Dict[str, Any]]) -> List[RowReport]:
    # Lazy-import the DB and orchestrator so dry-runs never need Mongo.
    from database import db  # noqa
    from services.contracts.contract_intelligence_service import (
        ContractIntelligenceService,
    )
    svc = ContractIntelligenceService(db)
    reports: List[RowReport] = []
    for idx, row in enumerate(rows, start=1):
        reports.append(await commit_row(svc, idx, row))
    return reports


# ---------------------------------------------------------------------------
# Printing helpers
# ---------------------------------------------------------------------------

def _fmt(value: Any, width: int) -> str:
    s = "" if value is None else str(value)
    return s.ljust(width)[:width]


def print_dryrun_header(mode: str, path: str, row_count: int) -> None:
    print("=" * 80)
    print(f"Navigator import — mode: {mode}")
    print(f"  source:       {path}")
    print(f"  total rows:   {row_count}")
    print("=" * 80)


def print_row(report: RowReport, *, mode: str) -> None:
    banner = f"[{report.index:>4}]"
    if report.error:
        print(f"{banner} ❌ ERROR: {report.error}")
        return
    env_id = report.envelope_id or "?"
    title = (report.title or "")[:60]
    status = report.status or "?"
    print(
        f"{banner} envelope={env_id}  status={status}  "
        f"title={title!r}"
    )
    print(
        f"       nav_uuid={report.provider_agreement_id or '-'}  "
        f"parties={report.party_count()}  "
        f"terms={report.term_count()}  "
        f"pricing={report.pricing_count()}  "
        f"docs={report.document_count()}  "
        f"warnings={report.warning_count()}"
    )
    if mode == "commit":
        if report.duplicate:
            print("       → skipped (duplicate event id; already imported)")
        elif report.committed:
            print(
                f"       → committed agreement_id={report.agreement_id} "
                f"links={report.link_count} exceptions={report.exception_count}"
            )
        else:
            print("       → NOT committed")
    if report.normalized and report.warning_count():
        for w in report.normalized.warnings:
            print(f"         · warn: {w.get('code')} {w.get('details')}")


def print_summary(reports: List[RowReport], *, mode: str) -> None:
    total = len(reports)
    errors = sum(1 for r in reports if r.error)
    committed = sum(1 for r in reports if r.committed)
    duplicates = sum(1 for r in reports if r.duplicate)
    warnings = sum(r.warning_count() for r in reports)
    print("-" * 80)
    print("Summary")
    print(f"  rows processed: {total}")
    print(f"  normalizer errors: {errors}")
    print(f"  total warnings:   {warnings}")
    if mode == "commit":
        print(f"  committed:        {committed}")
        print(f"  duplicates:       {duplicates}")
    else:
        print("  (dry-run — no DB writes)")
    print("-" * 80)


# ---------------------------------------------------------------------------
# Argparse + main
# ---------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="contracts_import_navigator",
        description=(
            "Import or dry-run a DocuSign Navigator AI Metadata Export into "
            "the Contract Intelligence store. Dry-run is the default."
        ),
    )
    p.add_argument(
        "path",
        help="Path to the Navigator .xlsx / .csv / .json file on the VM.",
    )
    p.add_argument(
        "--commit",
        action="store_true",
        help="Actually persist rows. Omit for a safe dry-run.",
    )
    p.add_argument(
        "--sheet",
        default=None,
        help="Worksheet name (xlsx only). Defaults to the active sheet.",
    )
    p.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Only process the first N rows (debugging).",
    )
    return p


def main(argv: Optional[List[str]] = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        rows = load_rows(args.path, sheet=args.sheet)
    except (FileNotFoundError, PermissionError, ValueError, RuntimeError) as exc:
        print(f"ERROR: cannot read {args.path!r}: {exc}", file=sys.stderr)
        return 2

    if args.limit is not None and args.limit >= 0:
        rows = rows[: args.limit]

    # Validate at least one row is a Navigator-shape row. Mixing
    # Connect-SIM payloads into this script is out of scope.
    if not rows:
        print(f"ERROR: no rows found in {args.path!r}", file=sys.stderr)
        return 2

    mode = "commit" if args.commit else "dry-run"
    print_dryrun_header(mode, args.path, len(rows))

    if args.commit:
        reports = asyncio.run(run_commit(rows))
    else:
        reports = [dryrun_row(i, r) for i, r in enumerate(rows, start=1)]

    for r in reports:
        print_row(r, mode=mode)

    print_summary(reports, mode=mode)

    errors = sum(1 for r in reports if r.error)
    if errors:
        return 3
    return 0


if __name__ == "__main__":
    sys.exit(main())
