"""Phase 4C(a) — Shared Navigator import service.

Single source of truth for parsing a DocuSign Navigator AI Metadata
Export and pushing the rows through the Contract Intelligence pipeline.
Used by:
  * ``scripts/contracts_import_navigator.py`` (CLI)
  * ``POST /api/contracts/navigator/import`` HTTP endpoint

The CLI and HTTP layers MUST NOT diverge — they both call:
  * :func:`parse_upload`            — bytes → list of row dicts
  * :func:`dryrun_rows`              — synchronous, no DB
  * :func:`commit_rows`              — async, persists via the orchestrator

Idempotency, audit, and matcher behavior are inherited from
:class:`ContractIntelligenceService` (deterministic event id
``navigator::{envelope_id}``; replays no-op at the
``agreement_events`` unique index).
"""

from __future__ import annotations

import csv
import io
import json
import logging
import os
from dataclasses import dataclass, field, asdict
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Limits
# ---------------------------------------------------------------------------

# Conservative default — Navigator exports are typically a few hundred KB.
# Override with ``CONTRACT_NAVIGATOR_IMPORT_MAX_BYTES``. Hard ceiling at
# 50 MB even if the env override is bigger; the orchestrator path is
# row-by-row and long-poll prone, so we want an absolute upper bound.
_DEFAULT_MAX_BYTES = 5 * 1024 * 1024  # 5 MB
_HARD_MAX_BYTES = 50 * 1024 * 1024    # 50 MB


def max_upload_bytes() -> int:
    raw = os.environ.get("CONTRACT_NAVIGATOR_IMPORT_MAX_BYTES")
    if not raw:
        return _DEFAULT_MAX_BYTES
    try:
        v = int(raw)
    except ValueError:
        return _DEFAULT_MAX_BYTES
    if v <= 0:
        return _DEFAULT_MAX_BYTES
    return min(v, _HARD_MAX_BYTES)


SUPPORTED_EXTENSIONS = (".xlsx", ".xlsm", ".csv", ".json")
SUPPORTED_CONTENT_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel.sheet.macroEnabled.12",
    "application/vnd.ms-excel",
    "text/csv",
    "application/csv",
    "application/json",
    "text/json",
    # Some browsers send octet-stream for xlsx — accepted, validated by
    # extension instead.
    "application/octet-stream",
}


class NavigatorImportError(ValueError):
    """Raised for any client-recoverable upload validation failure."""


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def _parse_xlsx_bytes(data: bytes, sheet: Optional[str]) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise NavigatorImportError(
            "openpyxl is required to read .xlsx files but is not installed"
        ) from exc
    bio = io.BytesIO(data)
    wb = load_workbook(filename=bio, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    if ws is None:
        return []
    rows: List[Dict[str, Any]] = []
    header_row: Optional[List[str]] = None
    for raw in ws.iter_rows(values_only=True):
        if header_row is None:
            header_row = [("" if c is None else str(c).strip()) for c in raw]
            continue
        if all((c is None or str(c).strip() == "") for c in raw):
            continue
        row_dict: Dict[str, Any] = {}
        for header, value in zip(header_row, raw):
            if not header:
                continue
            row_dict[header] = value
        rows.append(row_dict)
    return rows


def _parse_csv_bytes(data: bytes) -> List[Dict[str, Any]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(r) for r in reader]


def _parse_json_bytes(data: bytes) -> List[Dict[str, Any]]:
    payload = json.loads(data.decode("utf-8-sig", errors="replace"))
    if isinstance(payload, list):
        return [r for r in payload if isinstance(r, dict)]
    if isinstance(payload, dict):
        if isinstance(payload.get("row"), dict):
            return [payload["row"]]
        if isinstance(payload.get("rows"), list):
            return [r for r in payload["rows"] if isinstance(r, dict)]
        return [payload]
    raise NavigatorImportError("unsupported JSON shape; expected list/dict")


def parse_upload(
    *,
    data: bytes,
    filename: str,
    content_type: Optional[str] = None,
    sheet: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """Decode an uploaded Navigator file into a list of row dicts.

    Validation order: size cap → extension → content-type → parse.
    Raises :class:`NavigatorImportError` for any client-recoverable issue.
    """
    cap = max_upload_bytes()
    if len(data) > cap:
        raise NavigatorImportError(
            f"file is {len(data)} bytes; maximum allowed is {cap} bytes"
        )
    if not filename:
        raise NavigatorImportError("upload missing filename")
    ext = os.path.splitext(filename)[1].lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise NavigatorImportError(
            f"unsupported file extension {ext!r}; "
            f"allowed: {', '.join(SUPPORTED_EXTENSIONS)}"
        )
    if content_type and content_type.lower() not in SUPPORTED_CONTENT_TYPES:
        # Soft-warn rather than reject — extension is the canonical check.
        logger.info(
            "navigator import: unexpected content-type %r for %s",
            content_type, filename,
        )
    try:
        if ext in (".xlsx", ".xlsm"):
            rows = _parse_xlsx_bytes(data, sheet)
        elif ext == ".csv":
            rows = _parse_csv_bytes(data)
        else:  # .json
            rows = _parse_json_bytes(data)
    except NavigatorImportError:
        raise
    except (json.JSONDecodeError, UnicodeDecodeError) as exc:
        raise NavigatorImportError(f"file could not be parsed: {exc}") from exc
    except Exception as exc:  # noqa: BLE001
        raise NavigatorImportError(
            f"file could not be parsed: {type(exc).__name__}: {exc}"
        ) from exc
    if not rows:
        raise NavigatorImportError("no rows found in upload")
    return rows


# ---------------------------------------------------------------------------
# Per-row report + summary
# ---------------------------------------------------------------------------

@dataclass
class RowReport:
    index: int
    envelope_id: Optional[str] = None
    provider_agreement_id: Optional[str] = None
    title: Optional[str] = None
    status: Optional[str] = None

    party_count: int = 0
    term_count: int = 0
    pricing_count: int = 0
    document_count: int = 0
    warning_count: int = 0
    warnings: List[Dict[str, Any]] = field(default_factory=list)

    error: Optional[str] = None

    # Commit-only.
    committed: bool = False
    duplicate: bool = False
    agreement_id: Optional[str] = None
    link_count: int = 0
    exception_count: int = 0
    has_ambiguity_exception: bool = False

    # Phase 4C(d): per-row match preview (dry-run-only — empty in commit
    # mode, since the matcher already wrote the canonical link rows).
    match_preview: List[Dict[str, Any]] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


@dataclass
class ImportSummary:
    mode: str  # "dryrun" | "commit"
    filename: Optional[str] = None
    row_count: int = 0
    error_count: int = 0
    warning_count: int = 0

    # Aggregate counts (sum across rows that normalized successfully).
    agreements_detected: int = 0
    parties_detected: int = 0
    terms_detected: int = 0
    pricing_detected: int = 0
    documents_detected: int = 0

    # Commit-only.
    would_create: int = 0    # would-be-created agreements (dry-run)
    would_update: int = 0    # already-known envelope ids (dry-run)
    skipped: int = 0         # commit duplicates (idempotent re-import)
    committed: int = 0
    ambiguity_exceptions: int = 0
    schema_gap_warnings: int = 0

    # Phase 4C(d): BC matcher diagnostics.
    bc_repo_source: str = "empty"   # "reference_cache" | "empty"
    bc_cache_counts: Dict[str, int] = field(default_factory=dict)
    bc_cache_unavailable: bool = False

    rows: List[RowReport] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return {
            **{k: v for k, v in asdict(self).items() if k != "rows"},
            "rows": [r.to_dict() for r in self.rows],
        }


# ---------------------------------------------------------------------------
# Dry-run
# ---------------------------------------------------------------------------

def dryrun_one(index: int, row: Dict[str, Any]) -> RowReport:
    """Normalize a single row without writing anything."""
    # Local import — keeps the heavy normalizer modules lazy when consumers
    # only want the summary types.
    from services.contracts.navigator_normalizer import normalize_navigator_row

    report = RowReport(index=index)
    try:
        normalized = normalize_navigator_row(
            row, event_id=f"navigator-import::{index}",
        )
    except ValueError as exc:
        report.error = str(exc)
        return report
    except Exception as exc:  # noqa: BLE001
        report.error = f"{type(exc).__name__}: {exc}"
        return report

    report.envelope_id = normalized.agreement.provider_envelope_id
    report.provider_agreement_id = normalized.agreement.provider_agreement_id
    report.title = normalized.agreement.title
    report.status = normalized.agreement.status
    report.party_count = len(normalized.parties)
    report.term_count = len(normalized.terms)
    report.pricing_count = len(normalized.pricing)
    report.document_count = len(normalized.documents)
    report.warning_count = len(normalized.warnings)
    report.warnings = list(normalized.warnings)
    return report


async def _envelope_exists(db, envelope_id: str) -> bool:
    """Used in dry-run to compute would_create vs would_update."""
    if not envelope_id:
        return False
    from models.contracts import CONTRACTS_COLLECTIONS
    coll = db[CONTRACTS_COLLECTIONS["agreements"]]
    found = await coll.find_one(
        {"provider_envelope_id": envelope_id}, {"_id": 0, "id": 1},
    )
    return bool(found)


async def _build_bc_repo(db: Any) -> Tuple[Any, str, Dict[str, int], bool]:
    """Construct the BC repo to use for matching.

    Returns ``(repo, source, cache_counts, unavailable)``. When ``db`` is
    None or the cache probe fails, returns the in-memory empty repo so
    behavior degrades to today's "no candidates → exception" flow.
    """
    if db is None:
        from services.contracts.bc_agreement_matcher import InMemoryBCRepository
        return InMemoryBCRepository(), "empty", {}, False
    try:
        from services.contracts.bc_repo_reference_cache import (
            BCReferenceCacheRepository,
        )
        repo = BCReferenceCacheRepository(db)
        counts = await repo.probe()
        if repo.unavailable:
            from services.contracts.bc_agreement_matcher import InMemoryBCRepository
            return InMemoryBCRepository(), "empty", counts, True
        if not any(counts.values()):
            # Cache exists but cold — still use it (so future warm-up
            # transparently improves matches), but report "empty" source
            # so the UI can flag the cold-cache state.
            return repo, "reference_cache_cold", counts, False
        return repo, "reference_cache", counts, False
    except Exception as exc:  # noqa: BLE001
        logger.warning("falling back to empty BC repo: %s", exc)
        from services.contracts.bc_agreement_matcher import InMemoryBCRepository
        return InMemoryBCRepository(), "empty", {}, True


async def _preview_match_for_row(
    repo: Any, normalized: Any, *, ambiguity_band: float = 0.02,
    auto_confirm_threshold: float = 0.95, propose_threshold: float = 0.80,
) -> List[Dict[str, Any]]:
    """Run the matcher's lookup logic against the row's parties WITHOUT
    persisting links, so dry-run can show what would happen at commit.

    Returns a list of preview entries, one per party:
      ``{party_org, party_role, link_type, candidate_count,
         top_no, top_name, top_score, ambiguous, decision}``
    where ``decision`` is one of ``auto_confirm | propose | manual_review``.
    """
    out: List[Dict[str, Any]] = []
    for p in normalized.parties:
        link_type = "vendor" if (p.role or "").lower() == "sender" else "customer"
        if link_type == "customer":
            cands = await repo.find_customer_candidates(
                name=p.organization or p.name, email=p.email,
            )
        else:
            cands = await repo.find_vendor_candidates(
                name=p.organization or p.name, email=p.email,
            )
        entry: Dict[str, Any] = {
            "party_org": p.organization,
            "party_role": p.role,
            "link_type": link_type,
            "candidate_count": len(cands),
            "top_no": None,
            "top_name": None,
            "top_score": 0.0,
            "ambiguous": False,
            "decision": "manual_review",
        }
        if cands:
            top = cands[0]
            entry["top_no"] = top.no
            entry["top_name"] = top.name
            entry["top_score"] = top.score
            cluster = [
                c for c in cands
                if c.score >= propose_threshold
                and (top.score - c.score) <= ambiguity_band
                and c.no != top.no
            ]
            if cluster:
                entry["ambiguous"] = True
                entry["decision"] = "manual_review"
            elif top.score >= auto_confirm_threshold:
                entry["decision"] = "auto_confirm"
            elif top.score >= propose_threshold:
                entry["decision"] = "propose"
            else:
                entry["decision"] = "manual_review"
        out.append(entry)
    return out


async def dryrun_rows(
    rows: List[Dict[str, Any]],
    *,
    db: Any = None,
    filename: Optional[str] = None,
) -> ImportSummary:
    """Synchronous-friendly dry-run.

    If ``db`` is provided:
      * ``would_create`` / ``would_update`` is computed against the
        ``agreements`` collection.
      * Each row's ``match_preview`` is populated by running the BC
        repo against that row's parties (no persistence).
    """
    summary = ImportSummary(mode="dryrun", filename=filename, row_count=len(rows))

    repo, source, counts, unavailable = await _build_bc_repo(db)
    summary.bc_repo_source = source
    summary.bc_cache_counts = counts
    summary.bc_cache_unavailable = unavailable

    for idx, raw_row in enumerate(rows, start=1):
        report = dryrun_one(idx, raw_row)
        summary.rows.append(report)
        if report.error:
            summary.error_count += 1
            continue
        summary.agreements_detected += 1
        summary.parties_detected += report.party_count
        summary.terms_detected += report.term_count
        summary.pricing_detected += report.pricing_count
        summary.documents_detected += report.document_count
        summary.warning_count += report.warning_count
        for w in report.warnings:
            if (w.get("code") or "").startswith("schema_gap"):
                summary.schema_gap_warnings += 1
        if db is not None and report.envelope_id:
            exists = await _envelope_exists(db, report.envelope_id)
            if exists:
                summary.would_update += 1
            else:
                summary.would_create += 1
        else:
            summary.would_create += 1
        # Per-row match preview — re-normalize once for the matcher input.
        try:
            from services.contracts.navigator_normalizer import (
                normalize_navigator_row,
            )
            normalized = normalize_navigator_row(raw_row)
            report.match_preview = await _preview_match_for_row(repo, normalized)
        except Exception as exc:  # noqa: BLE001
            logger.warning("match preview failed for row %s: %s", idx, exc)
    return summary


# ---------------------------------------------------------------------------
# Commit
# ---------------------------------------------------------------------------

async def commit_one(svc, index: int, row: Dict[str, Any]) -> RowReport:
    """Push a single row through ``record_event`` + ``process_event``."""
    report = dryrun_one(index, row)
    if report.error or not report.envelope_id:
        return report

    event_id_str = f"navigator::{report.envelope_id}"
    record = await svc.record_event(
        provider_event_id=event_id_str,
        provider_envelope_id=report.envelope_id,
        event_type="navigator-import",
        raw_payload=row,
        hmac_valid=True,
        transport="manual",
    )
    if record["duplicate"]:
        report.duplicate = True
        return report

    processed = await svc.process_event(record["event_id"])
    if processed.get("status") == "ok":
        report.committed = True
    elif processed.get("status") == "normalizer_failed":
        report.error = processed.get("error") or "normalizer_failed"
        return report
    elif processed.get("status") == "post_normalize_failed":
        report.error = processed.get("error") or "post_normalize_failed"
        report.agreement_id = processed.get("agreement_id")
        return report
    report.agreement_id = processed.get("agreement_id")
    report.link_count = processed.get("links", 0) or 0
    report.exception_count = processed.get("exceptions", 0) or 0

    # Detect ambiguity exceptions on this agreement so the caller can
    # surface them in the response without a second round-trip. Guarded
    # against fakes/orchestrators that do not expose a real ``db``.
    if report.agreement_id and report.exception_count and getattr(svc, "db", None) is not None:
        from models.contracts import CONTRACTS_COLLECTIONS
        ex_coll = svc.db[CONTRACTS_COLLECTIONS["agreement_exceptions"]]
        cursor = ex_coll.find(
            {
                "agreement_id": report.agreement_id,
                "details.ambiguous": True,
            },
            {"_id": 0, "id": 1},
        )
        ambiguity = await cursor.to_list(length=1)
        report.has_ambiguity_exception = bool(ambiguity)
    return report


async def commit_rows(
    rows: List[Dict[str, Any]],
    *,
    db: Any,
    filename: Optional[str] = None,
) -> ImportSummary:
    """Write every row through the contract-intelligence orchestrator.

    Idempotency is delegated to ``ContractIntelligenceService.record_event``
    (unique index on ``(provider, provider_event_id)`` →
    ``navigator::{envelope_id}`` collapses replays).

    Phase 4C(d): the orchestrator is constructed with a
    :class:`BCReferenceCacheRepository` so the matcher can produce real
    candidates / auto-confirms / ambiguity exceptions instead of always
    returning empty.
    """
    from services.contracts.contract_intelligence_service import (
        ContractIntelligenceService,
    )

    summary = ImportSummary(mode="commit", filename=filename, row_count=len(rows))

    repo, source, counts, unavailable = await _build_bc_repo(db)
    summary.bc_repo_source = source
    summary.bc_cache_counts = counts
    summary.bc_cache_unavailable = unavailable

    svc = ContractIntelligenceService(db, repo=repo)
    for idx, raw_row in enumerate(rows, start=1):
        report = await commit_one(svc, idx, raw_row)
        summary.rows.append(report)
        if report.error:
            summary.error_count += 1
            continue
        summary.agreements_detected += 1
        summary.parties_detected += report.party_count
        summary.terms_detected += report.term_count
        summary.pricing_detected += report.pricing_count
        summary.documents_detected += report.document_count
        summary.warning_count += report.warning_count
        for w in report.warnings:
            if (w.get("code") or "").startswith("schema_gap"):
                summary.schema_gap_warnings += 1
        if report.duplicate:
            summary.skipped += 1
        if report.committed:
            summary.committed += 1
        if report.has_ambiguity_exception:
            summary.ambiguity_exceptions += 1
    return summary
