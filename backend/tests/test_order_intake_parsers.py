"""Phase 0 tests for deterministic customer order intake parsers."""

from datetime import datetime

from openpyxl import Workbook

from order_intake.parsers import CanPackXlsxParser, GiovanniOorParser, GiovanniQuantityProfile


def test_canpack_parser_normalizes_known_row(tmp_path):
    path = tmp_path / "canpack.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "september "
    ws.append([
        "", "", "", "PO#", "", "", "", "", "",
        "Customer Design Description",
        "Customer material number",
        "Call-off quantity",
        "Unit of Measurement (UoM)",
        "Delivering Plant",
        "Customer expected pick up date and time",
        "",
        "customer requested receive by date",
    ])
    ws.append([
        "", "", "", "W113159", "", "", "", "", "",
        "SPOTTED COW", "3286_NH01", 194500, "TS", "US50",
        datetime(2026, 9, 27, 8, 0), "", datetime(2026, 9, 29),
    ])
    wb.save(path)

    result = CanPackXlsxParser().parse(path)

    assert result.source.source_format == "CANPACK_XLSX"
    assert len(result.releases) == 1
    release = result.releases[0]
    assert release.customer_release_reference == "W113159"
    assert release.customer_item_reference == "3286_NH01"
    assert release.quantity == 194500
    assert release.uom == "TS"
    assert release.requested_delivery_date.isoformat() == "2026-09-29"


def test_giovanni_parser_preserves_packout_without_guessing_bc_uom(tmp_path):
    path = tmp_path / "gio.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    ws.cell(1, 8).value = "April 24oz Salsa"
    ws.cell(2, 8).value = 1
    ws.cell(2, 9).value = None
    ws.cell(2, 10).value = 61309
    ws.cell(2, 11).value = datetime(2026, 4, 23)
    ws.cell(2, 12).value = "Montreal? Run starts 4/20"

    # Existing Gamer PO must be excluded by default.
    ws.cell(3, 8).value = 2
    ws.cell(3, 9).value = 112999
    ws.cell(3, 10).value = 61310
    ws.cell(3, 11).value = datetime(2026, 4, 23)

    # SS is not a normal release reference.
    ws.cell(4, 8).value = 3
    ws.cell(4, 10).value = "SS"
    ws.cell(4, 11).value = datetime(2026, 4, 24)
    wb.save(path)

    parser = GiovanniOorParser(
        quantity_profiles={
            "24oz Salsa": GiovanniQuantityProfile(
                product_context="24oz Salsa",
                physical_units_per_load=58240,
                customer_item_reference="C-8682-10001486 / 12013925",
                # BC transaction quantity/UOM intentionally unresolved.
                source="Giovanni workbook packout evidence",
            )
        }
    )
    result = parser.parse(path, period="2026-04")

    assert len(result.releases) == 1
    release = result.releases[0]
    assert release.customer_release_reference == "61309"
    assert release.load_number == 1
    assert release.product_context == "24oz Salsa"
    assert release.physical_quantity == 58240
    assert release.physical_uom == "EA/PER_LOAD"
    assert release.quantity is None
    assert release.uom is None
    assert "BC sales quantity/UOM not resolved" in " | ".join(release.parser_review_reasons)
    assert "Montreal" in release.notes
    assert "Ship-to/location needs BC resolution" in release.parser_review_reasons


def test_giovanni_14oz_pizza_profile_maps_packout_to_m_uom(tmp_path):
    """Live BC evidence: a normal 89,775-unit load is transacted as 89.775 M."""
    path = tmp_path / "gio_pizza.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.cell(1, 22).value = "April 14oz Pizza"
    ws.cell(2, 22).value = 1
    ws.cell(2, 24).value = 61738
    ws.cell(2, 25).value = datetime(2026, 4, 23)
    wb.save(path)

    parser = GiovanniOorParser(
        quantity_profiles={
            "14oz Pizza": GiovanniQuantityProfile(
                product_context="14oz Pizza",
                physical_units_per_load=89775,
                customer_item_reference="C-8479-10000229",
                bc_quantity_per_load=89.775,
                bc_uom="M",
                source="live PRE_GAMERDOCS Giovanni sales history",
            )
        }
    )
    release = parser.parse(path, period="2026-04").releases[0]

    assert release.physical_quantity == 89775
    assert release.quantity == 89.775
    assert release.uom == "M"
    assert release.quantity_resolution_method == "authoritative_customer_item_profile"


def test_giovanni_16oz_vinegar_profile_maps_live_bc_m_uom(tmp_path):
    """Live PRE_GAMERDOCS evidence: a normal 78,166-unit load is 78.166 M."""
    path = tmp_path / "gio_vinegar.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.cell(1, 15).value = "April 16oz Vinegar"
    ws.cell(2, 15).value = 6
    ws.cell(2, 17).value = 61369
    ws.cell(2, 18).value = datetime(2026, 4, 14)
    wb.save(path)

    parser = GiovanniOorParser(
        quantity_profiles={
            "16oz Vinegar": GiovanniQuantityProfile(
                product_context="16oz Vinegar",
                physical_units_per_load=78166,
                customer_item_reference="C-8808-12026443",
                bc_quantity_per_load=78.166,
                bc_uom="M",
                source="live PRE_GAMERDOCS Giovanni sales history",
            )
        }
    )
    release = parser.parse(path, period="2026-04").releases[0]

    assert release.physical_quantity == 78166
    assert release.quantity == 78.166
    assert release.uom == "M"
    assert release.quantity_resolution_method == "authoritative_customer_item_profile"


def test_giovanni_nonstandard_load_does_not_inherit_normal_quantity(tmp_path):
    path = tmp_path / "gio_mixed.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.cell(1, 15).value = "April 16oz Vinegar"
    ws.cell(2, 15).value = 5
    ws.cell(2, 16).value = None
    ws.cell(2, 17).value = 61368
    ws.cell(2, 18).value = datetime(2026, 4, 9)
    ws.cell(2, 19).value = "21 pallets 16oz and 1 pallet 32oz"
    wb.save(path)

    parser = GiovanniOorParser(
        quantity_profiles={
            "16oz Vinegar": GiovanniQuantityProfile(
                product_context="16oz Vinegar",
                physical_units_per_load=78166,
                customer_item_reference="C-8808-12026443",
                bc_quantity_per_load=78.166,
                bc_uom="M",
                source="live PRE_GAMERDOCS Giovanni sales history",
            )
        }
    )
    release = parser.parse(path, period="2026-04").releases[0]

    assert release.quantity is None
    assert release.uom is None
    assert release.physical_quantity is None
    assert release.parser_review_reasons


def test_giovanni_period_filter_ignores_historical_same_month(tmp_path):
    path = tmp_path / "gio_history.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    ws.cell(1, 1).value = "April 24oz Pasta"
    ws.cell(2, 1).value = 1
    ws.cell(2, 3).value = 50001
    ws.cell(2, 4).value = datetime(2025, 4, 10)

    ws.cell(10, 1).value = "April 24oz Pasta"
    ws.cell(11, 1).value = 1
    ws.cell(11, 3).value = 61334
    ws.cell(11, 4).value = datetime(2026, 4, 14)
    wb.save(path)

    parser = GiovanniOorParser(
        quantity_profiles={
            "24oz Pasta": GiovanniQuantityProfile(
                product_context="24oz Pasta",
                physical_units_per_load=56420,
            )
        }
    )
    result = parser.parse(path, period="2026-04")

    assert [release.customer_release_reference for release in result.releases] == ["61334"]
