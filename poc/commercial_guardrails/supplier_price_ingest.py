from __future__ import annotations

import csv
import re
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Mapping, Sequence


HEADER_ALIASES = {
    "supplier_name": {
        "supplier", "supplier name", "vendor", "vendor name", "manufacturer",
    },
    "supplier_item_no": {
        "supplier item", "supplier item no", "supplier item number", "vendor item",
        "vendor item no", "vendor item number", "supplier sku", "vendor sku", "sku",
        "part", "part no", "part number",
    },
    "gpi_item_no": {
        "gpi item", "gpi item no", "gpi item number", "gamer item", "gamer item no",
        "gamer item number", "bc item", "bc item no", "bc item number",
    },
    "description": {
        "description", "item description", "product description", "product", "item",
    },
    "current_cost": {
        "current cost", "current price", "old cost", "old price", "previous cost",
        "previous price", "existing cost", "existing price",
    },
    "new_cost": {
        "new cost", "new price", "revised cost", "revised price", "updated cost",
        "updated price", "effective cost", "effective price", "price",
    },
    "effective_date": {
        "effective date", "effective", "price effective date", "new price effective",
        "new price effective date", "start date",
    },
    "uom": {
        "uom", "unit of measure", "unit", "pricing uom", "price uom",
    },
    "tier_qty": {
        "tier qty", "tier quantity", "minimum quantity", "min qty", "break qty",
        "quantity break", "price break",
    },
    "freight_included": {
        "freight included", "freight", "delivered", "delivered price", "includes freight",
    },
    "currency": {"currency", "currency code"},
}


@dataclass(frozen=True)
class SupplierPriceChange:
    supplier_name: str
    supplier_item_no: str
    gpi_item_no: str
    description: str
    current_cost: float | None
    new_cost: float | None
    effective_date: date | None
    uom: str
    tier_qty: float | None
    freight_included: bool | None
    currency: str
    source_file: str
    source_sheet: str
    source_row: int
    status: str
    warnings: tuple[str, ...]

    @property
    def cost_change(self) -> float | None:
        if self.current_cost is None or self.new_cost is None:
            return None
        return self.new_cost - self.current_cost

    @property
    def cost_change_pct(self) -> float | None:
        if self.current_cost is None or self.new_cost is None or self.current_cost == 0:
            return None
        return ((self.new_cost - self.current_cost) / abs(self.current_cost)) * 100.0

    def to_dict(self) -> dict:
        data = asdict(self)
        data["effective_date"] = self.effective_date.isoformat() if self.effective_date else ""
        data["warnings"] = "; ".join(self.warnings)
        data["cost_change"] = self.cost_change
        data["cost_change_pct"] = self.cost_change_pct
        return data


def _key(value: object) -> str:
    text = str(value or "").strip().casefold()
    text = re.sub(r"[_\-/]+", " ", text)
    text = re.sub(r"[^a-z0-9 ]+", "", text)
    return re.sub(r"\s+", " ", text).strip()


_ALIAS_LOOKUP = {
    _key(alias): canonical
    for canonical, aliases in HEADER_ALIASES.items()
    for alias in aliases
}


def canonical_header(value: object) -> str:
    return _ALIAS_LOOKUP.get(_key(value), "")


def map_headers(headers: Sequence[object]) -> dict[int, str]:
    mapped: dict[int, str] = {}
    for index, value in enumerate(headers):
        canonical = canonical_header(value)
        if canonical and canonical not in mapped.values():
            mapped[index] = canonical
    return mapped


def _text(value: object) -> str:
    return str(value if value is not None else "").strip()


def _float_or_none(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = _text(value)
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    text = text.replace("$", "").replace(",", "").replace("%", "")
    text = re.sub(r"\s*(USD|CAD|EUR)$", "", text, flags=re.IGNORECASE)
    try:
        number = float(text)
    except ValueError:
        return None
    return -number if negative else number


def _date_or_none(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    for fmt in (
        "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d",
        "%b %d %Y", "%B %d %Y", "%b %d, %Y", "%B %d, %Y",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _bool_or_none(value: object) -> bool | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    text = _key(value)
    if text in {"yes", "y", "true", "included", "include", "delivered", "1"}:
        return True
    if text in {"no", "n", "false", "excluded", "exclude", "not included", "0"}:
        return False
    return None


def _row_is_empty(values: Iterable[object]) -> bool:
    return not any(_text(value) for value in values)


def _stage_row(
    row: Mapping[str, object],
    *,
    default_supplier: str,
    source_file: str,
    source_sheet: str,
    source_row: int,
) -> SupplierPriceChange:
    supplier_name = _text(row.get("supplier_name")) or default_supplier.strip()
    supplier_item_no = _text(row.get("supplier_item_no"))
    gpi_item_no = _text(row.get("gpi_item_no"))
    description = _text(row.get("description"))
    current_cost = _float_or_none(row.get("current_cost"))
    new_cost = _float_or_none(row.get("new_cost"))
    effective_date = _date_or_none(row.get("effective_date"))
    uom = _text(row.get("uom")).upper()
    tier_qty = _float_or_none(row.get("tier_qty"))
    freight_included = _bool_or_none(row.get("freight_included"))
    currency = (_text(row.get("currency")) or "USD").upper()

    warnings: list[str] = []
    status = "READY"

    if not supplier_item_no and not gpi_item_no:
        warnings.append("missing supplier/GPI item identifier")
        status = "REJECT"
    if new_cost is None or new_cost <= 0:
        warnings.append("missing or invalid new cost")
        status = "REJECT"
    if not supplier_name:
        warnings.append("supplier name missing")
        if status != "REJECT":
            status = "REVIEW"
    if effective_date is None:
        warnings.append("effective date missing or unrecognized")
        if status != "REJECT":
            status = "REVIEW"
    if current_cost is None:
        warnings.append("current cost not supplied; BC comparison required")
        if status != "REJECT":
            status = "REVIEW"
    elif current_cost <= 0:
        warnings.append("current cost is zero or negative")
        if status != "REJECT":
            status = "REVIEW"
    if not uom:
        warnings.append("UOM missing; verify before comparing costs")
        if status != "REJECT":
            status = "REVIEW"

    return SupplierPriceChange(
        supplier_name=supplier_name,
        supplier_item_no=supplier_item_no,
        gpi_item_no=gpi_item_no,
        description=description,
        current_cost=current_cost,
        new_cost=new_cost,
        effective_date=effective_date,
        uom=uom,
        tier_qty=tier_qty,
        freight_included=freight_included,
        currency=currency,
        source_file=source_file,
        source_sheet=source_sheet,
        source_row=source_row,
        status=status,
        warnings=tuple(warnings),
    )


def _rows_from_csv(path: Path) -> tuple[str, list[tuple[int, dict[str, object]]]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        raw = list(reader)
    if not raw:
        return "", []
    header_map = map_headers(raw[0])
    if not header_map:
        raise ValueError("No recognized supplier-price headers were found in the CSV.")
    rows: list[tuple[int, dict[str, object]]] = []
    for row_number, values in enumerate(raw[1:], start=2):
        if _row_is_empty(values):
            continue
        mapped = {
            canonical: values[index] if index < len(values) else ""
            for index, canonical in header_map.items()
        }
        rows.append((row_number, mapped))
    return "", rows


def _rows_from_xlsx(path: Path, sheet_name: str = "") -> tuple[str, list[tuple[int, dict[str, object]]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX ingestion requires openpyxl. Install poc/commercial_guardrails/requirements.txt.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Worksheet {sheet_name!r} was not found. Available: {', '.join(workbook.sheetnames)}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]

        iterator = sheet.iter_rows(values_only=True)
        try:
            headers = next(iterator)
        except StopIteration:
            return sheet.title, []
        header_map = map_headers(headers)
        if not header_map:
            raise ValueError(f"No recognized supplier-price headers were found in worksheet {sheet.title!r}.")

        rows: list[tuple[int, dict[str, object]]] = []
        for row_number, values in enumerate(iterator, start=2):
            if _row_is_empty(values):
                continue
            mapped = {
                canonical: values[index] if index < len(values) else ""
                for index, canonical in header_map.items()
            }
            rows.append((row_number, mapped))
        return sheet.title, rows
    finally:
        workbook.close()


def load_supplier_notice(
    path: str | Path,
    *,
    default_supplier: str = "",
    sheet_name: str = "",
) -> list[SupplierPriceChange]:
    source = Path(path)
    suffix = source.suffix.casefold()
    if suffix == ".csv":
        source_sheet, raw_rows = _rows_from_csv(source)
    elif suffix in {".xlsx", ".xlsm"}:
        source_sheet, raw_rows = _rows_from_xlsx(source, sheet_name=sheet_name)
    else:
        raise ValueError(
            f"Unsupported supplier-price file type {source.suffix!r}. Current POC supports CSV and XLSX/XLSM."
        )

    return [
        _stage_row(
            row,
            default_supplier=default_supplier,
            source_file=source.name,
            source_sheet=source_sheet,
            source_row=row_number,
        )
        for row_number, row in raw_rows
    ]


def summarize_staging(rows: Sequence[SupplierPriceChange]) -> dict:
    ready = sum(row.status == "READY" for row in rows)
    review = sum(row.status == "REVIEW" for row in rows)
    reject = sum(row.status == "REJECT" for row in rows)
    increases = [row.cost_change_pct for row in rows if row.cost_change_pct is not None and row.cost_change_pct > 0]
    decreases = [row.cost_change_pct for row in rows if row.cost_change_pct is not None and row.cost_change_pct < 0]
    return {
        "rows": len(rows),
        "ready": ready,
        "review": review,
        "reject": reject,
        "increases": len(increases),
        "decreases": len(decreases),
        "max_increase_pct": max(increases) if increases else None,
        "max_decrease_pct": min(decreases) if decreases else None,
    }


def write_staging_csv(rows: Sequence[SupplierPriceChange], path: str | Path) -> None:
    fieldnames = [
        "supplier_name", "supplier_item_no", "gpi_item_no", "description",
        "current_cost", "new_cost", "cost_change", "cost_change_pct", "effective_date",
        "uom", "tier_qty", "freight_included", "currency", "status", "warnings",
        "source_file", "source_sheet", "source_row",
    ]
    with Path(path).open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row.to_dict())
